
import streamlit as st
import pandas as pd
import numpy as np
import sqlite3
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import io, os, warnings
warnings.filterwarnings("ignore")

st.set_page_config(page_title="APMI PMS Dashboard · Scripbox", page_icon="📊",
                   layout="wide", initial_sidebar_state="expanded")

st._config.set_option("theme.base", "light")
st.markdown("""
<style>
.stApp{background:#f8fafc}
[data-testid="stAppViewContainer"]{background:#f8fafc}
.stTabs [data-baseweb="tab-list"]{background:transparent}
.stTabs [data-baseweb="tab"]{color:inherit !important}
.stTabs [aria-selected="true"]{color:inherit !important; font-weight:600}
.sh{background:linear-gradient(135deg,#1e40af,#3b82f6);color:white;padding:12px 20px;
    border-radius:8px;margin:16px 0 12px;font-size:16px;font-weight:600}
.ib{background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;padding:12px 16px;margin:8px 0;font-size:14px;color:#1e40af}
.wb{background:#fffbeb;border:1px solid #fde68a;border-radius:8px;padding:12px 16px;margin:8px 0;font-size:14px;color:#92400e}
.sig-rec{background:#16a34a;color:white;border-radius:4px;padding:2px 8px;font-size:11px;font-weight:700}
.sig-hold{background:#d97706;color:white;border-radius:4px;padding:2px 8px;font-size:11px;font-weight:700}
.sig-nr{background:#dc2626;color:white;border-radius:4px;padding:2px 8px;font-size:11px;font-weight:700}
footer{visibility:hidden}
</style>""", unsafe_allow_html=True)

DB = "/tmp/apmi_pms.db"
COLORS = {"Equity":"#2563eb","Debt":"#7c3aed","Hybrid":"#059669","Multi Asset":"#d97706","benchmark":"#f59e0b","Inflow":"#16a34a"}
RET = {"return_1m":"1M","return_3m":"3M","return_6m":"6M","return_1y":"1Y","return_3y":"3Y","return_5y":"5Y","return_si":"SI"}
CAGR = {"return_1y","return_3y","return_5y","return_si"}
P2B  = {"return_1m":("bench_1m","alpha_1m"),"return_3m":("bench_3m","alpha_3m"),
        "return_6m":("bench_6m","alpha_6m"),"return_1y":("bench_1y","alpha_1y"),
        "return_3y":("bench_3y","alpha_3y"),"return_5y":("bench_5y","alpha_5y")}
FLAGS = {"CLEAN":"✅ Clean","WOUND_DOWN":"⚰️ Wound Down","STRUCTURED_PRODUCT":"🧾 Structured",
         "NEGATIVE_AUM_ERROR":"❌ Neg AUM","APMI_REPORTING_ERROR":"🚨 Reporting Error","APMI_REPORTED_ANOMALY":"⚠️ Anomaly"}
RF = 0.5417

# ── DB ────────────────────────────────────────────────────────────────────────
@st.cache_resource
def get_conn():
    return sqlite3.connect(DB, check_same_thread=False)
C = get_conn()

# ── DATA LOADERS ──────────────────────────────────────────────────────────────
@st.cache_data(ttl=3600)
def perf_raw():
    df = pd.read_sql("SELECT * FROM raw_performance", C)
    df["snapshot_date"] = pd.to_datetime(df["snapshot_date"])
    df = df.loc[:, ~df.columns.duplicated()]
    df["data_quality_flag"] = df["data_quality_flag"].fillna("CLEAN")
    return df

@st.cache_data(ttl=3600)
def bench_data():
    df = pd.read_sql("SELECT * FROM raw_benchmarks", C)
    df["snapshot_date"] = pd.to_datetime(df["snapshot_date"])
    return df

@st.cache_data(ttl=3600)
def risk_data():
    return pd.read_sql("SELECT * FROM risk_metrics", C)

@st.cache_data(ttl=3600)
def aum_decomp():
    df = pd.read_sql("SELECT * FROM aum_decomposition", C)
    df["snapshot_date"] = pd.to_datetime(df["snapshot_date"])
    return df

@st.cache_data(ttl=3600)
def turnover_data():
    df = pd.read_sql("SELECT * FROM ia_turnover", C)
    df["snapshot_date"] = pd.to_datetime(df["snapshot_date"])
    return df

def perf_with_bench():
    df = perf_raw(); db = bench_data()
    if db.empty: return df
    n = (db[db.benchmark_name.str.contains("Nifty 50", na=False)]
         [["snapshot_date","return_1m","return_3m","return_6m","return_1y","return_3y","return_5y"]]
         .rename(columns=lambda c: "bench_"+c[7:] if c != "snapshot_date" else c)
         .groupby("snapshot_date").first().reset_index())
    df = df.merge(n, on="snapshot_date", how="left")
    for p in ["1m","3m","6m","1y","3y","5y"]:
        if f"return_{p}" in df.columns and f"bench_{p}" in df.columns:
            df[f"alpha_{p}"] = df[f"return_{p}"] - df[f"bench_{p}"]
    return df

# ── AUTO-SIGNAL ENGINE ────────────────────────────────────────────────────────
# For each IA: compute rolling 18-month consistency score.
# Each month: did the IA finish in top half (Q1/Q2) vs strategy peers
# on BOTH absolute return AND Nifty 50 outperformance?
# Score = count of months where both conditions true / total months available.
# Recommended: score >= 0.78 (14/18), Hold: 0.50-0.77, Not Recommended: <0.50
@st.cache_data(ttl=3600)
def compute_signals():
    df = perf_raw()
    nifty = bench_data()
    if df.empty: return pd.DataFrame()

    clean = df[df["data_quality_flag"] == "CLEAN"].copy()
    nifty_m = (nifty[nifty.benchmark_name.str.contains("Nifty 50", na=False)]
               [["snapshot_date","return_1m"]].rename(columns={"return_1m":"nifty_1m"}))

    # Use last 18 months
    all_dates = sorted(clean["snapshot_date"].unique(), reverse=True)
    window = all_dates[:18]
    clean = clean[clean["snapshot_date"].isin(window)]
    clean = clean.merge(nifty_m, on="snapshot_date", how="left")
    clean["alpha_1m"] = clean["return_1m"] - clean["nifty_1m"].fillna(0)

    results = []
    for (prov, ia, svc, strat), grp in clean.groupby(
            ["pms_provider","ia_name","service_type","strategy_type"]):
        grp = grp.sort_values("snapshot_date")
        months_available = len(grp)
        if months_available < 6:
            continue

        top_half_both = 0
        for _, row in grp.iterrows():
            peers = clean[(clean["snapshot_date"] == row["snapshot_date"]) &
                          (clean["strategy_type"] == strat)]
            if len(peers) < 4:
                continue
            # Rank on absolute return
            abs_pct = (peers["return_1m"].dropna() <= row["return_1m"]).sum() / len(peers["return_1m"].dropna()) if pd.notna(row.get("return_1m")) else 0
            # Rank on Nifty outperformance
            alp_pct = (peers["alpha_1m"].dropna() <= row["alpha_1m"]).sum() / len(peers["alpha_1m"].dropna()) if pd.notna(row.get("alpha_1m")) else 0
            # Top half = percentile >= 0.50
            if abs_pct >= 0.50 and alp_pct >= 0.50:
                top_half_both += 1

        score = top_half_both / months_available
        if score >= 0.78:
            signal = "Recommended"
        elif score >= 0.50:
            signal = "Hold"
        else:
            signal = "Not Recommended"

        results.append({
            "pms_provider": prov, "ia_name": ia,
            "service_type": svc, "strategy_type": strat,
            "consistency_score": round(score, 3),
            "top_half_months": top_half_both,
            "months_available": months_available,
            "signal": signal,
        })

    return pd.DataFrame(results)

def sig_badge(signal):
    cls = {"Recommended":"sig-rec","Hold":"sig-hold","Not Recommended":"sig-nr"}.get(signal,"sig-hold")
    icons = {"Recommended":"✅","Hold":"⏸️","Not Recommended":"❌"}
    return f"<span class='{cls}'>{icons.get(signal,'')} {signal}</span>"

# ── HELPERS ───────────────────────────────────────────────────────────────────
def snap(df, s):
    if s: return df[df.snapshot_date.dt.strftime("%Y-%m-%d") == s]
    return df[df.snapshot_date == df.snapshot_date.max()]

def fp(v, cagr=False):
    if v is None or (isinstance(v, float) and np.isnan(v)): return "—"
    return f"{float(v):+.2f}{'% CAGR' if cagr else '%'}"

def excel_single_pms(pms_provider, ia_name, service_type, snap_date):
    """Generate multi-sheet Excel for a single PMS — one sheet per return period."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io as _io

        # Load full history for this IA
        df_hist = pd.read_sql("""
            SELECT snapshot_date, return_1m, return_3m, return_6m,
                   return_1y, return_3y, return_5y, return_si,
                   aum_cr, data_quality_flag
            FROM raw_performance
            WHERE pms_provider=? AND ia_name=? AND service_type=?
            ORDER BY snapshot_date DESC
        """, C, params=[pms_provider, ia_name, service_type])

        # Load benchmarks for same dates
        db_hist = pd.read_sql("""
            SELECT snapshot_date, benchmark_name,
                   return_1m, return_3m, return_6m,
                   return_1y, return_3y, return_5y
            FROM raw_benchmarks
            ORDER BY snapshot_date DESC
        """, C)

        df_hist["snapshot_date"] = pd.to_datetime(df_hist["snapshot_date"])
        db_hist["snapshot_date"] = pd.to_datetime(db_hist["snapshot_date"])

        PERIODS = {
            "1M":  ("return_1m",  False),
            "3M":  ("return_3m",  False),
            "6M":  ("return_6m",  False),
            "1Y":  ("return_1y",  True),
            "3Y":  ("return_3y",  True),
            "5Y":  ("return_5y",  True),
            "SI":  ("return_si",  True),
        }

        BENCH_NAMES = ["Nifty 50 TRI", "BSE 500 TRI", "Nifty Midcap 150 TRI"]

        # Styles
        BLUE      = "1E40AF"
        GREEN     = "16A34A"
        RED       = "DC2626"
        LGREY     = "F1F5F9"
        WHITE     = "FFFFFF"
        hdr_fill  = PatternFill("solid", fgColor=BLUE)
        hdr_font  = Font(color=WHITE, bold=True, size=10)
        grn_fill  = PatternFill("solid", fgColor="D1FAE5")
        red_fill  = PatternFill("solid", fgColor="FEE2E2")
        grn_font  = Font(color=GREEN, bold=True)
        red_font  = Font(color=RED,   bold=True)
        alt_fill  = PatternFill("solid", fgColor=LGREY)
        center    = Alignment(horizontal="center", vertical="center")
        left      = Alignment(horizontal="left",   vertical="center")
        thin      = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0"),
        )

        output = _io.BytesIO()
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default sheet

        # ── SUMMARY SHEET (all periods side by side) ──────────────────────
        ws_sum = wb.create_sheet("Summary — All Periods")
        ws_sum.freeze_panes = "A2"

        sum_headers = ["Snapshot Date", "AUM (₹ Cr)", "Flag",
                       "1M %", "3M %", "6M %",
                       "1Y % (CAGR)", "3Y % (CAGR)", "5Y % (CAGR)", "Since Inception"]
        bench_headers = []
        for bn in BENCH_NAMES:
            for p_label in ["1M", "3M", "6M", "1Y", "3Y", "5Y"]:
                bench_headers.append(f"{bn} {p_label}")

        all_headers = sum_headers + bench_headers
        ws_sum.append(all_headers)

        # Style summary header
        for col_i, _ in enumerate(all_headers, 1):
            cell = ws_sum.cell(row=1, column=col_i)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = center; cell.border = thin

        # Summary data rows
        for row_i, (_, row) in enumerate(df_hist.iterrows(), 2):
            snap_str = row["snapshot_date"].strftime("%b %Y")
            values = [
                snap_str,
                round(row["aum_cr"], 1) if pd.notna(row.get("aum_cr")) else "—",
                row.get("data_quality_flag") or "CLEAN",
                round(row["return_1m"], 2)  if pd.notna(row.get("return_1m")) else "—",
                round(row["return_3m"], 2)  if pd.notna(row.get("return_3m")) else "—",
                round(row["return_6m"], 2)  if pd.notna(row.get("return_6m")) else "—",
                round(row["return_1y"], 2)  if pd.notna(row.get("return_1y")) else "—",
                round(row["return_3y"], 2)  if pd.notna(row.get("return_3y")) else "—",
                round(row["return_5y"], 2)  if pd.notna(row.get("return_5y")) else "—",
                round(row["return_si"], 2)  if pd.notna(row.get("return_si")) else "—",
            ]
            # Add benchmark values for this date
            db_snap = db_hist[db_hist["snapshot_date"] == row["snapshot_date"]]
            for bn in BENCH_NAMES:
                bn_row = db_snap[db_snap["benchmark_name"] == bn]
                for p_col in ["return_1m","return_3m","return_6m","return_1y","return_3y","return_5y"]:
                    if not bn_row.empty and pd.notna(bn_row[p_col].iloc[0]):
                        values.append(round(float(bn_row[p_col].iloc[0]), 2))
                    else:
                        values.append("—")
            ws_sum.append(values)
            fill = alt_fill if row_i % 2 == 0 else PatternFill()
            for col_i in range(1, len(all_headers)+1):
                cell = ws_sum.cell(row=row_i, column=col_i)
                cell.alignment = center; cell.border = thin
                if col_i > 3:  # return columns
                    try:
                        v = float(cell.value)
                        cell.fill = grn_fill if v > 0 else (red_fill if v < 0 else fill)
                        cell.font = grn_font if v > 0 else (red_font if v < 0 else Font(size=10))
                    except (TypeError, ValueError):
                        cell.fill = fill
                else:
                    cell.fill = fill

        # Auto-width summary
        for col_i, header in enumerate(all_headers, 1):
            ws_sum.column_dimensions[get_column_letter(col_i)].width = max(len(str(header))+2, 10)
        ws_sum.row_dimensions[1].height = 30

        # ── ONE SHEET PER PERIOD ───────────────────────────────────────────
        for p_label, (p_col, is_cagr) in PERIODS.items():
            sheet_name = f"{p_label}{' (CAGR)' if is_cagr else ''}"
            ws = wb.create_sheet(sheet_name)
            ws.freeze_panes = "A2"

            headers = ["Snapshot Date", "AUM (₹ Cr)",
                       f"{ia_name[:30]} Return",
                       "Nifty 50 TRI", "BSE 500 TRI", "Nifty Midcap 150 TRI",
                       "Alpha vs Nifty 50", "vs Category (placeholder)"]
            ws.append(headers)
            for col_i, _ in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_i)
                cell.fill = hdr_fill; cell.font = hdr_font
                cell.alignment = center; cell.border = thin

            for row_i, (_, row) in enumerate(df_hist.iterrows(), 2):
                if p_col not in row or pd.isna(row.get(p_col)):
                    pms_ret = None
                else:
                    pms_ret = round(float(row[p_col]), 2)

                db_snap = db_hist[db_hist["snapshot_date"] == row["snapshot_date"]]
                bench_vals = {}
                for bn in BENCH_NAMES:
                    bn_row = db_snap[db_snap["benchmark_name"] == bn]
                    bv = round(float(bn_row[p_col].iloc[0]), 2) if (
                        not bn_row.empty and p_col in bn_row.columns
                        and pd.notna(bn_row[p_col].iloc[0])
                    ) else None
                    bench_vals[bn] = bv

                nifty_val = bench_vals.get("Nifty 50 TRI")
                alpha = round(pms_ret - nifty_val, 2) if (
                    pms_ret is not None and nifty_val is not None) else None

                row_vals = [
                    row["snapshot_date"].strftime("%b %Y"),
                    round(row["aum_cr"], 1) if pd.notna(row.get("aum_cr")) else "—",
                    pms_ret if pms_ret is not None else "—",
                    bench_vals.get("Nifty 50 TRI") or "—",
                    bench_vals.get("BSE 500 TRI") or "—",
                    bench_vals.get("Nifty Midcap 150 TRI") or "—",
                    alpha if alpha is not None else "—",
                    "—",
                ]
                ws.append(row_vals)

                row_fill = alt_fill if row_i % 2 == 0 else PatternFill()
                for col_i in range(1, len(headers)+1):
                    cell = ws.cell(row=row_i, column=col_i)
                    cell.alignment = center; cell.border = thin
                    if col_i >= 3:
                        try:
                            v = float(cell.value)
                            cell.fill = grn_fill if v > 0 else (red_fill if v < 0 else row_fill)
                            cell.font = grn_font if v > 0 else (red_font if v < 0 else Font(size=10))
                        except (TypeError, ValueError):
                            cell.fill = row_fill
                    else:
                        cell.fill = row_fill

            # Alpha conditional format (col 7)
            for row_i in range(2, len(df_hist)+2):
                cell = ws.cell(row=row_i, column=7)
                try:
                    v = float(cell.value)
                    cell.fill = grn_fill if v > 0 else (red_fill if v < 0 else PatternFill())
                    cell.font = Font(color=GREEN if v > 0 else RED, bold=True)
                except (TypeError, ValueError):
                    pass

            for col_i, h in enumerate(headers, 1):
                ws.column_dimensions[get_column_letter(col_i)].width = max(len(str(h))+2, 12)
            ws.row_dimensions[1].height = 30

        # ── METADATA SHEET ─────────────────────────────────────────────────
        ws_meta = wb.create_sheet("About")
        meta_rows = [
            ["APMI PMS Research Export"],
            [""],
            ["PMS Provider",    pms_provider],
            ["Strategy Name",   ia_name],
            ["Service Type",    service_type],
            ["Snapshot Date",   snap_date or "Latest"],
            ["Generated",       pd.Timestamp.now().strftime("%d %b %Y %H:%M")],
            [""],
            ["Data Source",     "apmiindia.org — SEBI-mandated TWRR"],
            ["Returns >1Y",     "Shown as CAGR"],
            ["Alpha",           "PMS return minus Nifty 50 TRI return"],
            ["Disclaimer",      "Internal research use only. Not investment advice."],
        ]
        for r in meta_rows:
            ws_meta.append(r)
        ws_meta.column_dimensions["A"].width = 20
        ws_meta.column_dimensions["B"].width = 60
        ws_meta["A1"].font = Font(bold=True, size=14, color=BLUE)

        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        import traceback
        st.error(f"Excel export error: {e}")
        st.code(traceback.format_exc())
        return None

def sec(t): st.markdown(f"<div class='sh'>{t}</div>", unsafe_allow_html=True)
def exp(t):
    with st.expander("💡 What does this mean?"): st.markdown(f"<div class='ib'>{t}</div>", unsafe_allow_html=True)
def csvb(df, fn, lbl="⬇️ CSV"): st.download_button(lbl, df.to_csv(index=False).encode(), fn, "text/csv")

# ── SIDEBAR ───────────────────────────────────────────────────────────────────
def sidebar(df):
    with st.sidebar:
        st.markdown("## 📊 APMI PMS Dashboard")
        st.markdown("*Scripbox Internal Research*"); st.markdown("---")
        strats = sorted(df.strategy_type.dropna().unique()) if not df.empty else []
        ss = st.multiselect("Strategy", strats, default=strats)
        svcs = sorted(df.service_type.dropna().unique()) if not df.empty else []
        sv = st.multiselect("Service Type", svcs, default=svcs)
        pl = st.selectbox("Return Period", list(RET.values()), index=3)
        p = [k for k,v in RET.items() if v == pl][0]
        ma = st.slider("Min AUM (₹ Cr)", 0, 1000, 0, 50)
        ps = st.text_input("Search Provider", "")
        dates = sorted(df.snapshot_date.dt.strftime("%Y-%m-%d").unique(), reverse=True) if not df.empty else []
        sd = st.selectbox("Snapshot Date", dates) if dates else None
        st.markdown("---"); st.markdown("**Data Quality**")
        sf = st.toggle("Include flagged rows", value=False)
        fl = ["CLEAN"]
        if sf:
            av = sorted(df.data_quality_flag.dropna().unique()) if not df.empty else ["CLEAN"]
            fl = st.multiselect("Flags", av, default=av, format_func=lambda f: FLAGS.get(f, f))
            if not fl: fl = ["CLEAN"]
        else:
            st.caption("CLEAN rows only.")
        st.markdown("---")
        sig_filter = st.selectbox("Signal Filter", ["All","Recommended","Hold","Not Recommended"], index=0)
        if not df.empty:
            st.markdown("---")
            st.caption(f"Latest: **{df.snapshot_date.max().strftime('%b %Y')}**")
            st.caption(f"**{len(df):,}** records · apmiindia.org")
    return dict(strategies=ss, service_types=sv, period=p, period_label=pl,
                min_aum=ma, provider_search=ps, snapshot_date=sd,
                selected_flags=fl, show_flagged=sf, sig_filter=sig_filter)

def apply_filters(df, f, signals=None):
    if df.empty: return df
    df = snap(df, f.get("snapshot_date"))
    if f["strategies"]: df = df[df.strategy_type.isin(f["strategies"])]
    if f["service_types"]: df = df[df.service_type.isin(f["service_types"])]
    if f["min_aum"] > 0: df = df[df.aum_cr >= f["min_aum"]]
    if f["provider_search"]: df = df[df.pms_provider.str.lower().str.contains(f["provider_search"].lower(), na=False)]
    if "data_quality_flag" in df.columns and f.get("selected_flags"):
        df = df[df.data_quality_flag.isin(f["selected_flags"])]
    # Merge signals and filter
    if signals is not None and not signals.empty and f.get("sig_filter","All") != "All":
        df = df.merge(signals[["pms_provider","ia_name","service_type","signal"]], on=["pms_provider","ia_name","service_type"], how="left")
        df = df[df["signal"] == f["sig_filter"]]
        df = df.drop(columns=["signal"], errors="ignore")
    return df

# ── TAB 1: LEADERBOARD ────────────────────────────────────────────────────────
def t_leaderboard(df, f, db, signals):
    sec("🏆 TWRR Leaderboard")
    exp("Ranked by TWRR. Signal = auto-computed from last 18 months of rolling returns vs strategy peers. "
        "Score = % of months in top half on BOTH absolute return and Nifty 50 outperformance. "
        "Recommended ≥78%, Hold 50-77%, Not Recommended <50%.")
    if df.empty: st.info("No data."); return
    p, pl = f["period"], f["period_label"]
    df = df.loc[:, ~df.columns.duplicated()]
    if p not in df.columns: st.warning(f"{pl} not available."); return

    bc, ac = P2B.get(p, (None, None))
    base = [c for c in ["pms_provider","ia_name","strategy_type","service_type","aum_cr","data_quality_flag"] if c in df.columns]
    ex = [c for c in [bc, ac] if c and c in df.columns]
    lb = (df[base + [p] + ex].dropna(subset=[p]).sort_values(p, ascending=False).copy())
    lb.insert(0, "Rank", range(1, len(lb)+1))

    # Delta vs last month — query DB directly, independent of all filters
    try:
        snaps_d = pd.read_sql(
            "SELECT DISTINCT snapshot_date FROM raw_performance ORDER BY snapshot_date DESC LIMIT 2",
            C)["snapshot_date"].tolist()
        if len(snaps_d) >= 2:
            curr_snap, prev_snap = snaps_d[0], snaps_d[1]
            # Current full ranking (unfiltered, same period)
            curr_sql = f"SELECT pms_provider, ia_name, {p} as curr_ret FROM raw_performance WHERE snapshot_date=? AND (data_quality_flag IS NULL OR data_quality_flag='CLEAN') AND {p} IS NOT NULL"
            curr_full = pd.read_sql(curr_sql, C, params=[curr_snap]).sort_values("curr_ret", ascending=False).reset_index(drop=True)
            curr_full["curr_rank"] = curr_full.index + 1
            prev_sql = f"SELECT pms_provider, ia_name, {p} as prev_ret FROM raw_performance WHERE snapshot_date=? AND (data_quality_flag IS NULL OR data_quality_flag='CLEAN') AND {p} IS NOT NULL"
            prev_full = pd.read_sql(prev_sql, C, params=[prev_snap]).sort_values("prev_ret", ascending=False).reset_index(drop=True)
            prev_full["prev_rank"] = prev_full.index + 1
            # Merge both into lb
            lb = lb.merge(curr_full[["pms_provider","ia_name","curr_rank"]],
                          on=["pms_provider","ia_name"], how="left")
            lb = lb.merge(prev_full[["pms_provider","ia_name","prev_rank","prev_ret"]],
                          on=["pms_provider","ia_name"], how="left")
            lb["Δ Rank"] = lb.apply(lambda r:
                "new" if pd.isna(r.get("prev_rank")) else
                (f"▲{int(r['prev_rank']-r['curr_rank'])}" if pd.notna(r.get("curr_rank")) and int(r["prev_rank"]-r["curr_rank"]) > 0 else
                (f"▼{int(abs(r['prev_rank']-r['curr_rank']))}" if pd.notna(r.get("curr_rank")) and int(r["prev_rank"]-r["curr_rank"]) < 0
                 else "━")), axis=1)
            lb["Δ Return"] = lb.apply(lambda r:
                "new" if pd.isna(r.get("prev_ret")) else
                f"{float(r[p])-float(r['prev_ret']):+.2f}%", axis=1)
        else:
            lb["Δ Rank"] = "—"; lb["Δ Return"] = "—"
    except Exception as e:
        lb["Δ Rank"] = "—"; lb["Δ Return"] = "—"
        st.warning(f"Delta debug: {type(e).__name__}: {e}")

    # Merge signals
    if not signals.empty:
        lb = lb.merge(signals[["pms_provider","ia_name","service_type","signal","consistency_score","top_half_months","months_available"]],
                      on=["pms_provider","ia_name","service_type"], how="left")
        lb["Signal"] = lb["signal"].fillna("—")
        lb["Score"] = lb.apply(lambda r: f"{r['top_half_months']:.0f}/{r['months_available']:.0f}" if pd.notna(r.get("top_half_months")) else "—", axis=1)
    else:
        lb["Signal"] = "—"; lb["Score"] = "—"

    # Quartile within strategy
    def aq(g):
        g = g.copy()
        try: g["Q"] = pd.qcut(g[p], q=4, labels=["Q4","Q3","Q2","Q1"], duplicates="drop")
        except: g["Q"] = "N/A"
        return g
    lb = lb.groupby("strategy_type", group_keys=False).apply(aq)

    ic = p in CAGR
    c1,c2,c3,c4 = st.columns(4)
    c1.metric("IAs shown", f"{len(lb):,}")
    c2.metric(f"Best ({pl})", fp(lb[p].max(), ic))
    c3.metric(f"Worst ({pl})", fp(lb[p].min(), ic))
    c4.metric(f"Median ({pl})", fp(lb[p].median(), ic))

    if ac and ac in lb.columns:
        pb = (lb[ac] > 0).mean() * 100
        bv = lb[bc].dropna().iloc[0] if bc and bc in lb.columns and not lb[bc].dropna().empty else None
        b1, b2 = st.columns(2)
        if bv is not None: b1.metric(f"Nifty 50 ({pl})", fp(bv, ic))
        b2.metric("% Beat Nifty", f"{pb:.1f}%")

    # Signal summary
    if "signal" in lb.columns:
        sc = lb["signal"].value_counts()
        s1, s2, s3 = st.columns(3)
        s1.metric("✅ Recommended", sc.get("Recommended", 0))
        s2.metric("⏸️ Hold", sc.get("Hold", 0))
        s3.metric("❌ Not Recommended", sc.get("Not Recommended", 0))

    st.markdown("---")
    rn = {"pms_provider":"Provider","ia_name":"IA","strategy_type":"Strategy",
          "service_type":"Service","aum_cr":"AUM(₹Cr)","data_quality_flag":"Flag",
          p:f"Return({pl})"}
    if bc and bc in lb.columns: rn[bc] = f"Nifty50({pl})"
    if ac and ac in lb.columns: rn[ac] = "Alpha"
    dd = lb.rename(columns=rn)
    if "Δ Rank" not in dd.columns and "Δ Rank" in lb.columns: dd["Δ Rank"] = lb["Δ Rank"].values
    if "Δ Return" not in dd.columns and "Δ Return" in lb.columns: dd["Δ Return"] = lb["Δ Return"].values
    # Format Nifty50 and Alpha as percentages
    for _fc in [f"Nifty50({pl})", "Alpha"]:
        if _fc in dd.columns:
            dd[_fc] = pd.to_numeric(dd[_fc], errors="coerce").apply(
                lambda x: f"{x:+.2f}%" if pd.notna(x) else "—")
    # Also format Return column
    ret_col_name = f"Return({pl})"
    if ret_col_name in dd.columns:
        dd[ret_col_name] = pd.to_numeric(dd[ret_col_name], errors="coerce").apply(
            lambda x: f"{x:+.2f}%" if pd.notna(x) else "—")
    sc2 = (["Rank","Δ Rank","Δ Return","Signal","Score","Provider","IA","Strategy","AUM(₹Cr)",f"Return({pl})"]
           + ([f"Nifty50({pl})"] if f"Nifty50({pl})" in dd.columns else [])
           + (["Alpha"] if "Alpha" in dd.columns else [])
           + ["Q","Flag"])
    sc2 = [c for c in sc2 if c in dd.columns]
    st.dataframe(dd[sc2], use_container_width=True, height=520, hide_index=True)
    csvb(dd[sc2], f"leaderboard_{pl}.csv", f"⬇️ CSV — {pl}")

    # Distribution chart
    fig = px.histogram(lb, x=p, nbins=50, color="strategy_type" if "strategy_type" in lb.columns else None, color_discrete_map=COLORS,
                       title=f"Distribution of {pl} Returns")
    fig.add_vline(x=lb[p].median(), line_dash="dash", annotation_text=f"Median: {lb[p].median():.2f}%")
    if bc and bc in lb.columns:
        bv2 = lb[bc].dropna()
        if not bv2.empty:
            fig.add_vline(x=bv2.iloc[0], line_dash="dot", line_color="#f59e0b",
                          annotation_text=f"Nifty50: {bv2.iloc[0]:.2f}%", annotation_font_color="#f59e0b")
    fig.update_layout(height=300, template="plotly_white")
    st.plotly_chart(fig, use_container_width=True)

# ── TAB 2: SIGNALS ────────────────────────────────────────────────────────────
def t_signals(signals, f, df_raw, db):
    sec("🎯 Auto-Signals — Consistency Ranking")
    exp("Signal computed from last 18 months of rolling 1M returns vs strategy peers. "
        "Each month: was the PMS in the top half on BOTH absolute return AND Nifty 50 outperformance? "
        "Score = months both conditions true ÷ total months available. "
        "Recommended ≥14/18 (78%), Hold 9-13/18 (50-77%), Not Recommended <9/18. "
        "Within Recommended, higher score = more consistent = better.")
    if signals.empty: st.info("Not enough data to compute signals."); return

    strats = f.get("strategies", [])
    sig = signals[signals.strategy_type.isin(strats)] if strats else signals.copy()

    # Filter by signal type
    sc1, sc2 = st.columns(2)
    with sc1:
        sel_strat_sig = st.selectbox("Filter by Strategy", ["All"] + sorted(sig.strategy_type.unique().tolist()), key="sig_strat")
    with sc2:
        sel_sig = st.radio("Show", ["All","Recommended","Hold","Not Recommended"], horizontal=True, key="sig_radio")
    if sel_strat_sig != "All": sig = sig[sig["strategy_type"] == sel_strat_sig]
    if sel_sig != "All": sig = sig[sig["signal"] == sel_sig]

    # Summary metrics
    total = len(signals[signals.strategy_type.isin(strats)] if strats else signals)
    rec = len(signals[(signals.strategy_type.isin(strats) if strats else True) & (signals["signal"]=="Recommended")])
    hold = len(signals[(signals.strategy_type.isin(strats) if strats else True) & (signals["signal"]=="Hold")])
    nr = len(signals[(signals.strategy_type.isin(strats) if strats else True) & (signals["signal"]=="Not Recommended")])

    m1,m2,m3,m4 = st.columns(4)
    m1.metric("Total IAs scored", total)
    m2.metric("✅ Recommended", rec, help="≥14/18 months top-half on both lenses")
    m3.metric("⏸️ Hold", hold, help="9–13/18 months")
    m4.metric("❌ Not Recommended", nr, help="<9/18 months")

    st.markdown("---")

    # Signal breakdown by strategy
    c1, c2 = st.columns(2)
    with c1:
        sb = (signals[signals.strategy_type.isin(strats)] if strats else signals).groupby(["strategy_type","signal"]).size().reset_index(name="count")
        fig = px.bar(sb, x="strategy_type", y="count", color="signal",
                     color_discrete_map={"Recommended":"#16a34a","Hold":"#d97706","Not Recommended":"#dc2626"},
                     barmode="stack", title="Signal breakdown by strategy")
        fig.update_layout(height=320, template="plotly_white", legend=dict(orientation="h", y=-0.25))
        st.plotly_chart(fig, use_container_width=True)
    with c2:
        # Score distribution
        fig2 = px.histogram(signals[signals.strategy_type.isin(strats)] if strats else signals,
                            x="consistency_score", nbins=20, color="signal",
                            color_discrete_map={"Recommended":"#16a34a","Hold":"#d97706","Not Recommended":"#dc2626"},
                            title="Score distribution (0=never top-half, 1=always top-half)")
        fig2.add_vline(x=0.78, line_dash="dash", line_color="#16a34a", annotation_text="Rec threshold")
        fig2.add_vline(x=0.50, line_dash="dash", line_color="#d97706", annotation_text="Hold threshold")
        fig2.update_layout(height=320, template="plotly_white", legend=dict(orientation="h", y=-0.25))
        st.plotly_chart(fig2, use_container_width=True)

    # Main table sorted by score desc
    disp = sig[["pms_provider","ia_name","strategy_type","service_type",
                "signal","consistency_score","top_half_months","months_available"]].copy()
    disp["Score"] = disp.apply(lambda r: f"{int(r.top_half_months)}/{int(r.months_available)}", axis=1)
    disp["Score %"] = (disp["consistency_score"] * 100).round(1)
    disp = disp.sort_values("consistency_score", ascending=False).rename(columns={
        "pms_provider":"Provider","ia_name":"IA","strategy_type":"Strategy",
        "service_type":"Service","signal":"Signal"})
    show = ["Signal","Score","Score %","Provider","IA","Strategy","Service"]
    st.dataframe(disp[show], use_container_width=True, height=520, hide_index=True)
    csvb(disp[show], "signals.csv", "⬇️ CSV — All signals")

    # ── CONSISTENCY HEATMAP ────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### 🔥 Consistency Heatmap — month-by-month Q ranking")
    st.caption("Each cell = one month. Colour = quartile rank vs strategy peers on absolute return. "
               "Green=Q1 (top), yellow=Q2, orange=Q3, red=Q4, grey=no data. "
               "Read left to right = oldest to newest. The pattern matters more than the score.")

    hm_strat = st.selectbox("Strategy", ["All"] + sorted(sig.strategy_type.unique().tolist()), key="hm_strat")
    hm_sig   = st.radio("Signal", ["Recommended","Hold","Not Recommended","All"], horizontal=True, key="hm_sig_r")

    hm_df = sig.copy()
    if hm_strat != "All": hm_df = hm_df[hm_df.strategy_type == hm_strat]
    if hm_sig != "All":   hm_df = hm_df[hm_df.signal == hm_sig]
    hm_df = hm_df.sort_values("consistency_score", ascending=False).head(50)

    if hm_df.empty:
        st.info("No IAs match filter.")
    else:
        all_dates_hm = sorted(perf_raw()["snapshot_date"].unique(), reverse=False)[-18:]
        month_labels  = [pd.Timestamp(d).strftime("%b%y") for d in all_dates_hm]

        nifty_hm = (bench_data()[bench_data().benchmark_name.str.contains("Nifty 50",na=False)]
                    [["snapshot_date","return_1m"]].rename(columns={"return_1m":"nifty_1m"}))

        clean_all = perf_raw()[perf_raw().data_quality_flag=="CLEAN"].copy()
        clean_all = clean_all.merge(nifty_hm, on="snapshot_date", how="left")
        clean_all["alpha_1m"] = clean_all["return_1m"] - clean_all["nifty_1m"].fillna(0)

        Q_COLORS = {1:"#16a34a", 2:"#86efac", 3:"#fb923c", 4:"#dc2626", None:"#e5e7eb"}
        Q_LABELS = {1:"Q1", 2:"Q2", 3:"Q3", 4:"Q4", None:"—"}

        def get_q(val, series):
            s = series.dropna()
            if len(s) < 4 or pd.isna(val): return None
            pct = (s <= val).sum() / len(s)
            return 1 if pct >= 0.75 else (2 if pct >= 0.50 else (3 if pct >= 0.25 else 4))

        # Build HTML table
        col_w = 38
        html = f"""
        <style>
        .hm-table{{border-collapse:collapse;font-size:11px;width:100%;table-layout:fixed}}
        .hm-table th{{background:#1e40af;color:white;padding:4px 2px;text-align:center;font-size:10px;position:sticky;top:0}}
        .hm-table td{{padding:3px 2px;text-align:center;border:1px solid #f1f5f9;font-size:10px}}
        .hm-ia{{text-align:left !important;padding-left:6px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;max-width:200px;font-weight:500}}
        .hm-score{{font-weight:700;color:#1e40af;background:#eff6ff}}
        </style>
        <div style="overflow-x:auto">
        <table class="hm-table">
        <thead><tr>
        <th style="text-align:left;min-width:200px">IA Name</th>
        <th style="min-width:60px">Signal</th>
        <th style="min-width:50px">Score</th>
        """
        for ml in month_labels:
            html += f'<th style="min-width:{col_w}px">{ml}</th>'
        html += "</tr></thead><tbody>"

        sig_colors = {"Recommended":"#16a34a","Hold":"#d97706","Not Recommended":"#dc2626"}

        for _, row_hm in hm_df.iterrows():
            ia_peers_all = clean_all[clean_all.strategy_type == row_hm.strategy_type]
            ia_hist_hm   = clean_all[(clean_all.pms_provider==row_hm.pms_provider)&
                                      (clean_all.ia_name==row_hm.ia_name)&
                                      (clean_all.service_type==row_hm.service_type)]
            sc = f"{int(row_hm.top_half_months)}/{int(row_hm.months_available)}"
            sc_label = row_hm.signal
            sc_color = sig_colors.get(sc_label, "#94a3b8")
            html += f"""<tr>
            <td class="hm-ia" title="{row_hm.ia_name}">{row_hm.ia_name[:40]}</td>
            <td style="color:white;background:{sc_color};font-weight:700;font-size:9px">{sc_label}</td>
            <td class="hm-score">{sc}</td>"""
            for dt in all_dates_hm:
                row_dt = ia_hist_hm[ia_hist_hm.snapshot_date == dt]
                peers_dt = ia_peers_all[ia_peers_all.snapshot_date == dt]
                if row_dt.empty or row_dt.return_1m.isna().all():
                    q = None
                else:
                    val = float(row_dt.return_1m.iloc[0])
                    q   = get_q(val, peers_dt["return_1m"])
                bg  = Q_COLORS[q]
                lbl = Q_LABELS[q]
                ret_str = f"{float(row_dt.return_1m.iloc[0]):+.1f}%" if not row_dt.empty and pd.notna(row_dt.return_1m.iloc[0] if not row_dt.empty else None) else ""
                html += f'<td style="background:{bg};color:{"white" if q in [1,4] else "#1e293b"}" title="{ret_str}">{lbl}</td>'
            html += "</tr>"

        html += "</tbody></table></div>"
        st.markdown(html, unsafe_allow_html=True)
        st.caption(f"Showing top {len(hm_df)} IAs by consistency score. Hover over a cell to see the actual return.")

    # Manual override
    st.markdown("---")
    with st.expander("✏️ Manual override — disagree with auto-signal?"):
        st.caption("Auto-signal is based on 18-month rolling consistency. Override only when you have context the data doesn't capture (e.g. fund manager change, strategy pivot).")
        try:
            C.execute("""CREATE TABLE IF NOT EXISTS overrides(
                pms_provider TEXT, ia_name TEXT, service_type TEXT,
                override_signal TEXT, reason TEXT,
                UNIQUE(pms_provider,ia_name,service_type))""")
            C.commit()
        except: pass
        ov_opts = signals[["pms_provider","ia_name","service_type","signal"]].copy()
        ov_labels = [f"{r.pms_provider} — {r.ia_name} ({r.service_type}) [{r.signal}]" for r in ov_opts.itertuples()]
        if ov_labels:
            ov_sel = st.selectbox("Select IA to override", ov_labels, key="ov_sel")
            ov_row = ov_opts.iloc[ov_labels.index(ov_sel)]
            ov_sig = st.radio("Override signal", ["Recommended","Hold","Not Recommended"], horizontal=True, key="ov_sig")
            ov_reason = st.text_input("Reason (required)", key="ov_reason", placeholder="e.g. Fund manager changed in Jan 2026")
            if st.button("Save override", key="ov_save"):
                if ov_reason.strip():
                    C.execute("""INSERT INTO overrides(pms_provider,ia_name,service_type,override_signal,reason)
                        VALUES(?,?,?,?,?) ON CONFLICT(pms_provider,ia_name,service_type)
                        DO UPDATE SET override_signal=excluded.override_signal,reason=excluded.reason""",
                        (ov_row.pms_provider,ov_row.ia_name,ov_row.service_type,ov_sig,ov_reason.strip()))
                    C.commit(); st.cache_data.clear(); st.success(f"✅ Override saved: {ov_row.ia_name} → {ov_sig}"); st.rerun()
                else:
                    st.warning("Please enter a reason before saving.")
        # Show existing overrides
        try:
            ov_df = pd.read_sql("SELECT * FROM overrides", C)
            if not ov_df.empty:
                st.markdown("**Current overrides:**")
                st.dataframe(ov_df.rename(columns={"pms_provider":"Provider","ia_name":"IA",
                    "service_type":"Service","override_signal":"Override","reason":"Reason"}),
                    use_container_width=True, hide_index=True)
        except: pass

# ── TAB 3: BENCHMARK COMPARE ──────────────────────────────────────────────────
def t_bench(df_raw, f, db, signals):
    sec("📐 Benchmark Comparison")
    exp("Point-in-time returns vs Nifty 50 TRI, BSE 500 TRI, Nifty Midcap 150 TRI + category median. Returns >1Y are CAGR.")
    if df_raw.empty: st.info("No data."); return
    s = f.get("snapshot_date")
    cl = snap(df_raw, s); cl = cl[cl.data_quality_flag == "CLEAN"]
    il = cl[["pms_provider","ia_name","service_type"]].drop_duplicates().sort_values(["pms_provider","ia_name"])
    if il.empty: st.info("No clean data."); return
    all_provs_bc = sorted(il.pms_provider.unique().tolist())
    prov_search_bc = st.text_input("🔍 Search provider", "", key="bc_prov_search", placeholder="e.g. Motilal, HDFC, Sundaram...")
    filtered_provs_bc = [p2 for p2 in all_provs_bc if prov_search_bc.lower() in p2.lower()] if prov_search_bc else all_provs_bc
    if not filtered_provs_bc: st.info("No providers match."); return
    sel_prov_bc = st.selectbox("Select Provider", filtered_provs_bc, key="bc_prov_sel")
    ia_filtered_bc = il[il.pms_provider == sel_prov_bc]
    ia_opts_bc = [f"{r.ia_name} ({r.service_type})" for r in ia_filtered_bc.itertuples()]
    if not ia_opts_bc: st.info("No strategies for this provider."); return
    sel_ia_bc = st.selectbox("Select Strategy", ia_opts_bc, key="bc_ia_sel")
    sr = ia_filtered_bc.iloc[ia_opts_bc.index(sel_ia_bc)]
    pr = cl[(cl.pms_provider==sr.pms_provider)&(cl.ia_name==sr.ia_name)&(cl.service_type==sr.service_type)].iloc[0]
    pr_row = pr  # used by PDF generator
    bs = snap(db, s)

    # Show signal for this IA
    if not signals.empty:
        sig_row = signals[(signals.pms_provider==sr.pms_provider)&(signals.ia_name==sr.ia_name)&(signals.service_type==sr.service_type)]
        if not sig_row.empty:
            sr2 = sig_row.iloc[0]
            st.markdown(f"**Signal:** {sr2['signal']} &nbsp;|&nbsp; **Score:** {int(sr2.top_half_months)}/{int(sr2.months_available)} months top-half &nbsp;|&nbsp; **Strategy:** {sr2.strategy_type}", unsafe_allow_html=True)
            st.markdown("---")

    ps = ["1M","3M","6M","1Y","3Y","5Y","Since Inception"]
    rc = ["return_1m","return_3m","return_6m","return_1y","return_3y","return_5y","return_si"]
    rows = [{"Name":pr.ia_name,"Type":"PMS",**{p:fp(pr.get(c),c in CAGR) for p,c in zip(ps,rc)}}]
    for _, b in bs.iterrows():
        rows.append({"Name":b.benchmark_name,"Type":"Benchmark",**{p:fp(b.get(c),c in CAGR) for p,c in zip(ps,rc)}})
    peers = cl[cl.strategy_type == pr.strategy_type]
    rows.append({"Name":f"{pr.strategy_type} Median","Type":"Peer median",
                 **{p:fp(peers[c].median() if c in peers.columns else None,c in CAGR) for p,c in zip(ps,rc)}})
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    nr = bs[bs.benchmark_name.str.contains("Nifty 50", na=False)]
    if not nr.empty:
        st.markdown("#### Alpha vs Nifty 50 TRI")
        ad = []
        for p2, c in zip(ps[:6], rc[:6]):
            pv = pr.get(c); nv = nr[c].iloc[0] if c in nr.columns else None
            if pv is not None and nv is not None and pd.notna(pv) and pd.notna(nv):
                ad.append({"Period":p2,"Alpha(%)":round(float(pv)-float(nv),2)})
        if ad:
            fig = px.bar(pd.DataFrame(ad), x="Period", y="Alpha(%)",
                         color="Alpha(%)", color_continuous_scale=["#dc2626","#f8fafc","#16a34a"],
                         color_continuous_midpoint=0, title=f"Alpha vs Nifty 50 — {pr.ia_name}")
            fig.add_hline(y=0, line_dash="dash", line_color="#94a3b8")
            fig.update_layout(height=300, template="plotly_white", coloraxis_showscale=False)
            st.plotly_chart(fig, use_container_width=True)
    c1, c2 = st.columns(2)
    with c1:
        csvb(pd.DataFrame(rows), f"bench_{s or 'latest'}.csv")
    with c2:
        if st.button("📊 Download Full Excel (all periods)", key="xl_bench"):
            xl = excel_single_pms(sr.pms_provider, sr.ia_name, sr.service_type, s)
            if xl:
                ia_clean = sr.ia_name[:30].replace(' ','_').replace('/','_')
                snap_str = pd.Timestamp(s).strftime('%b_%Y') if s else 'latest'
                st.download_button(
                    "⬇️ Click to download Excel",
                    data=xl,
                    file_name=f"PMS_{ia_clean}_{snap_str}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="xl_dl"
                )

    # ── PDF REPORT ─────────────────────────────────────────────────────────────
    st.markdown("---")
    if st.button("📄 Generate PDF Report for this PMS", key="pdf_btn"):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
            import io as _io

            buf = _io.BytesIO()
            doc = SimpleDocTemplate(buf, pagesize=A4,
                                    rightMargin=2*cm, leftMargin=2*cm,
                                    topMargin=2*cm, bottomMargin=2*cm)
            styles = getSampleStyleSheet()
            story  = []

            # Colours
            BLUE  = colors.HexColor("#1e40af")
            GREEN = colors.HexColor("#16a34a")
            RED   = colors.HexColor("#dc2626")
            AMBER = colors.HexColor("#d97706")
            LGREY = colors.HexColor("#f1f5f9")
            DGREY = colors.HexColor("#64748b")

            # Custom styles
            title_style = ParagraphStyle("title", parent=styles["Heading1"],
                fontSize=20, textColor=BLUE, spaceAfter=4)
            sub_style   = ParagraphStyle("sub", parent=styles["Normal"],
                fontSize=10, textColor=DGREY, spaceAfter=12)
            h2_style    = ParagraphStyle("h2", parent=styles["Heading2"],
                fontSize=13, textColor=BLUE, spaceBefore=14, spaceAfter=6)
            body_style  = ParagraphStyle("body", parent=styles["Normal"],
                fontSize=9, spaceAfter=4)
            bold_style  = ParagraphStyle("bold", parent=styles["Normal"],
                fontSize=9, fontName="Helvetica-Bold")

            snap_str = pd.Timestamp(s).strftime("%B %Y") if s else perf_raw().snapshot_date.max().strftime("%B %Y")

            # ── PAGE 1 ────────────────────────────────────────────────────────
            story.append(Paragraph("APMI PMS Research Report", title_style))
            story.append(Paragraph(f"Scripbox Internal Research &nbsp;·&nbsp; {snap_str}", sub_style))
            story.append(HRFlowable(width="100%", thickness=2, color=BLUE))
            story.append(Spacer(1, 0.3*cm))

            # IA header
            ia_full = f"{pr_row['ia_name']} — {pr_row['pms_provider']}"
            story.append(Paragraph(ia_full, ParagraphStyle("iahdr", parent=styles["Heading1"],
                fontSize=16, textColor=colors.HexColor("#0f172a"), spaceAfter=4)))
            story.append(Paragraph(
                f"Strategy: {pr_row.get('strategy_type','—')} &nbsp;·&nbsp; "
                f"Service: {pr_row.get('service_type','—')} &nbsp;·&nbsp; "
                f"AUM: ₹{pr_row.get('aum_cr',0):,.0f} Cr", sub_style))
            story.append(Spacer(1, 0.2*cm))

            # Signal box
            sig_row = signals[(signals.pms_provider==sr.pms_provider)&
                               (signals.ia_name==sr.ia_name)&
                               (signals.service_type==sr.service_type)] if not signals.empty else pd.DataFrame()
            if not sig_row.empty:
                sr3 = sig_row.iloc[0]
                sig_color = {"Recommended":GREEN,"Hold":AMBER,"Not Recommended":RED}.get(sr3.signal, DGREY)
                sig_data = [[
                    Paragraph(f"<b>Signal: {sr3.signal}</b>", ParagraphStyle("sc", fontSize=11, textColor=colors.white)),
                    Paragraph(f"Consistency Score: {int(sr3.top_half_months)}/{int(sr3.months_available)} months ({sr3.consistency_score*100:.0f}%)",
                              ParagraphStyle("sc2", fontSize=10, textColor=colors.white)),
                    Paragraph(f"Strategy: {sr3.strategy_type}", ParagraphStyle("sc3", fontSize=10, textColor=colors.white)),
                ]]
                sig_table = Table(sig_data, colWidths=["30%","40%","30%"])
                sig_table.setStyle(TableStyle([
                    ("BACKGROUND",(0,0),(-1,-1), sig_color),
                    ("ROWBACKGROUNDS",(0,0),(-1,-1),[sig_color]),
                    ("TOPPADDING",(0,0),(-1,-1),8),
                    ("BOTTOMPADDING",(0,0),(-1,-1),8),
                    ("LEFTPADDING",(0,0),(-1,-1),10),
                    ("ROUNDEDCORNERS",[4]),
                ]))
                story.append(sig_table)
                story.append(Spacer(1, 0.4*cm))

            # Returns vs benchmarks table
            story.append(Paragraph("Point-in-time Returns vs Benchmarks", h2_style))
            story.append(Paragraph("All returns >1Y are CAGR. Source: APMI (SEBI-mandated TWRR).", body_style))

            ret_headers = ["Name","Type","1M","3M","6M","1Y","3Y","5Y","Since Inception"]
            table_data  = [ret_headers]
            for row_p in rows:
                table_data.append([
                    row_p.get("Name",""),
                    row_p.get("Type",""),
                    row_p.get("1M","—"),
                    row_p.get("3M","—"),
                    row_p.get("6M","—"),
                    row_p.get("1Y","—"),
                    row_p.get("3Y","—"),
                    row_p.get("5Y","—"),
                    row_p.get("Since Inception","—"),
                ])
            col_ws = [3.5*cm, 2.2*cm, 1.5*cm, 1.5*cm, 1.5*cm, 1.8*cm, 1.8*cm, 1.8*cm, 2.2*cm]
            ret_table = Table(table_data, colWidths=col_ws)
            ret_style = TableStyle([
                ("BACKGROUND",  (0,0),(-1,0),  BLUE),
                ("TEXTCOLOR",   (0,0),(-1,0),  colors.white),
                ("FONTNAME",    (0,0),(-1,0),  "Helvetica-Bold"),
                ("FONTSIZE",    (0,0),(-1,-1), 8),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, LGREY]),
                ("GRID",        (0,0),(-1,-1), 0.5, colors.HexColor("#e2e8f0")),
                ("TOPPADDING",  (0,0),(-1,-1), 5),
                ("BOTTOMPADDING",(0,0),(-1,-1),5),
                ("LEFTPADDING", (0,0),(-1,-1), 6),
                ("FONTNAME",    (0,1),(0,-1),  "Helvetica-Bold"),
            ])
            # Colour PMS return cells green/red
            for col_i, period in enumerate(["1M","3M","6M","1Y"], start=2):
                try:
                    val_str = table_data[1][col_i]
                    val_num = float(val_str.replace("%","").replace(" CAGR","").replace("+","").replace("—","nan"))
                    if not pd.isna(val_num):
                        cell_color = GREEN if val_num > 0 else RED
                        ret_style.add("TEXTCOLOR", (col_i,1),(col_i,1), cell_color)
                        ret_style.add("FONTNAME",  (col_i,1),(col_i,1), "Helvetica-Bold")
                except: pass
            ret_table.setStyle(ret_style)
            story.append(ret_table)
            story.append(Spacer(1, 0.5*cm))

            # Risk metrics
            story.append(Paragraph("Risk Metrics", h2_style))
            risk_df = risk_data()
            ia_risk = risk_df[(risk_df.pms_provider==sr.pms_provider)&
                               (risk_df.ia_name==sr.ia_name)&
                               (risk_df.service_type==sr.service_type)] if not risk_df.empty else pd.DataFrame()
            if not ia_risk.empty:
                ir2 = ia_risk.iloc[0]
                risk_tdata = [[
                    Paragraph("<b>Sharpe Ratio</b>",   bold_style),
                    Paragraph(f"{ir2.get('sharpe_ratio','—'):.2f}"  if pd.notna(ir2.get("sharpe_ratio"))  else "—", body_style),
                    Paragraph("<b>Sortino Ratio</b>",  bold_style),
                    Paragraph(f"{ir2.get('sortino_ratio','—'):.2f}" if pd.notna(ir2.get("sortino_ratio")) else "—", body_style),
                ],[
                    Paragraph("<b>Std Deviation</b>",  bold_style),
                    Paragraph(f"{ir2.get('std_deviation','—'):.2f}%" if pd.notna(ir2.get("std_deviation")) else "—", body_style),
                    Paragraph("<b>Info Ratio</b>",     bold_style),
                    Paragraph(f"{ir2.get('info_ratio','—'):.2f}"    if pd.notna(ir2.get("info_ratio"))    else "—", body_style),
                ],[
                    Paragraph("<b>Max Drawdown</b>",   bold_style),
                    Paragraph(f"{ir2.get('max_drawdown','—'):.2f}%" if pd.notna(ir2.get("max_drawdown")) else "—", body_style),
                    Paragraph("<b>Alpha vs Nifty</b>", bold_style),
                    Paragraph(f"{ir2.get('alpha_vs_nifty50','—'):.2f}%" if pd.notna(ir2.get("alpha_vs_nifty50")) else "—", body_style),
                ]]
                risk_table = Table(risk_tdata, colWidths=[3.5*cm,2.5*cm,3.5*cm,2.5*cm])
                risk_table.setStyle(TableStyle([
                    ("ROWBACKGROUNDS",(0,0),(-1,-1),[colors.white, LGREY]),
                    ("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#e2e8f0")),
                    ("TOPPADDING",(0,0),(-1,-1),6),
                    ("BOTTOMPADDING",(0,0),(-1,-1),6),
                    ("LEFTPADDING",(0,0),(-1,-1),8),
                ]))
                story.append(risk_table)
                story.append(Spacer(1, 0.3*cm))
                story.append(Paragraph(
                    f"Based on {int(ir2.get('months_of_data',0))} months of return data. "
                    "Risk-free rate: 6.5% p.a. (0.54% monthly). "
                    "Information Ratio >0.5 indicates strong manager skill vs benchmark.",
                    ParagraphStyle("note", parent=styles["Normal"],
                        fontSize=8, textColor=DGREY, fontName="Helvetica-Oblique")))

            # ── PAGE 2 ────────────────────────────────────────────────────────
            story.append(Spacer(1, 0.5*cm))
            story.append(HRFlowable(width="100%", thickness=1, color=LGREY))
            story.append(Spacer(1, 0.3*cm))
            story.append(Paragraph("Consistency Analysis — Last 18 Months", h2_style))
            story.append(Paragraph(
                "Each row shows the PMS's quartile rank vs strategy peers for that month on two lenses: "
                "absolute return (Abs) and outperformance of Nifty 50 TRI (vs Nifty). "
                "Q1 = top 25%. Q4 = bottom 25%.",
                body_style))
            story.append(Spacer(1, 0.2*cm))

            # Quartile history table
            hi_pdf = pd.read_sql("""
                SELECT snapshot_date, return_1m, return_1y FROM raw_performance
                WHERE pms_provider=? AND ia_name=? AND service_type=?
                AND (data_quality_flag IS NULL OR data_quality_flag='CLEAN')
                ORDER BY snapshot_date DESC LIMIT 18
            """, C, params=[sr.pms_provider, sr.ia_name, sr.service_type])
            pe_pdf = pd.read_sql("""
                SELECT snapshot_date, return_1m FROM raw_performance
                WHERE strategy_type=(SELECT strategy_type FROM raw_performance WHERE pms_provider=? AND ia_name=? LIMIT 1)
                AND (data_quality_flag IS NULL OR data_quality_flag='CLEAN')
            """, C, params=[sr.pms_provider, sr.ia_name])
            nf_pdf = pd.read_sql("""
                SELECT snapshot_date, return_1m AS n1m FROM raw_benchmarks
                WHERE benchmark_name='Nifty 50 TRI'
            """, C)
            for d in [hi_pdf, pe_pdf, nf_pdf]:
                d["snapshot_date"] = pd.to_datetime(d["snapshot_date"])

            def qr_pdf(v, s2):
                s2 = s2.dropna()
                if len(s2) < 4 or pd.isna(v): return "—"
                pct = (s2 <= v).sum() / len(s2)
                return "Q1" if pct >= 0.75 else ("Q2" if pct >= 0.50 else ("Q3" if pct >= 0.25 else "Q4"))

            q_header = ["Month","1M Return","Q-1M Abs","Q-1M Nifty","1Y Return","Q-1Y Abs"]
            q_tdata  = [q_header]
            for _, rq in hi_pdf.iterrows():
                ps_q = pe_pdf[pe_pdf.snapshot_date==rq.snapshot_date]
                ns_q = nf_pdf[nf_pdf.snapshot_date==rq.snapshot_date]
                n1m_q = float(ns_q.n1m.iloc[0]) if not ns_q.empty and pd.notna(ns_q.n1m.iloc[0]) else None
                alpha_q = (rq.return_1m - n1m_q) if n1m_q and pd.notna(rq.return_1m) else None
                peers_alpha = ps_q.copy(); peers_alpha["alpha"] = peers_alpha.return_1m - (n1m_q or 0)
                q_tdata.append([
                    rq.snapshot_date.strftime("%b %Y"),
                    f"{rq.return_1m:+.2f}%" if pd.notna(rq.return_1m) else "—",
                    qr_pdf(rq.return_1m, ps_q.return_1m),
                    qr_pdf(alpha_q, peers_alpha.alpha),
                    f"{rq.return_1y:+.2f}%" if pd.notna(rq.get("return_1y")) else "—",
                    qr_pdf(rq.get("return_1y"), ps_q.return_1m),
                ])
            q_table = Table(q_tdata, colWidths=[2.5*cm,2.5*cm,2.5*cm,2.5*cm,2.5*cm,2.5*cm])
            q_ts = TableStyle([
                ("BACKGROUND",      (0,0),(-1,0), BLUE),
                ("TEXTCOLOR",       (0,0),(-1,0), colors.white),
                ("FONTNAME",        (0,0),(-1,0), "Helvetica-Bold"),
                ("FONTSIZE",        (0,0),(-1,-1),8),
                ("ROWBACKGROUNDS",  (0,1),(-1,-1),[colors.white, LGREY]),
                ("GRID",            (0,0),(-1,-1),0.5, colors.HexColor("#e2e8f0")),
                ("TOPPADDING",      (0,0),(-1,-1),5),
                ("BOTTOMPADDING",   (0,0),(-1,-1),5),
                ("LEFTPADDING",     (0,0),(-1,-1),6),
            ])
            # Colour Q cells
            q_colors_pdf = {"Q1":GREEN,"Q2":colors.HexColor("#86efac"),
                            "Q3":colors.HexColor("#fb923c"),"Q4":RED}
            for row_i, row_q in enumerate(q_tdata[1:], start=1):
                for col_i in [2,3,5]:
                    qval = row_q[col_i]
                    if qval in q_colors_pdf:
                        q_ts.add("BACKGROUND", (col_i,row_i),(col_i,row_i), q_colors_pdf[qval])
                        q_ts.add("TEXTCOLOR",  (col_i,row_i),(col_i,row_i),
                                 colors.white if qval in ["Q1","Q4"] else colors.HexColor("#1e293b"))
                        q_ts.add("FONTNAME",   (col_i,row_i),(col_i,row_i), "Helvetica-Bold")
            q_table.setStyle(q_ts)
            story.append(q_table)
            story.append(Spacer(1, 0.5*cm))

            # Footer
            story.append(HRFlowable(width="100%", thickness=1, color=LGREY))
            story.append(Spacer(1, 0.2*cm))
            story.append(Paragraph(
                f"Generated by APMI PMS Dashboard · Scripbox Internal Research · {snap_str} · "
                "Data source: apmiindia.org (SEBI-mandated TWRR) · Not investment advice",
                ParagraphStyle("footer", parent=styles["Normal"],
                    fontSize=7, textColor=DGREY, alignment=TA_CENTER)))

            doc.build(story)
            buf.seek(0)

            ia_name_clean = sr.ia_name[:30].replace(" ","_").replace("/","_")
            st.download_button(
                label=f"⬇️ Download PDF — {sr.ia_name[:40]}",
                data=buf.getvalue(),
                file_name=f"PMS_Report_{ia_name_clean}_{snap_str.replace(' ','_')}.pdf",
                mime="application/pdf",
                key="pdf_download"
            )
            st.success("✅ PDF ready — click above to download")

        except ImportError:
            st.error("reportlab not installed. Run: !pip install reportlab -q")
        except Exception as e:
            st.error(f"PDF generation error: {e}")
            import traceback; st.code(traceback.format_exc())

# ── TAB 4: NAV GROWTH ─────────────────────────────────────────────────────────
def t_nav(df_raw, f, db, signals):
    sec("📈 NAV Growth vs Benchmark")
    exp("NAV indexed to 100. Info Ratio = active return ÷ tracking error vs Nifty 50. Above 0.5 = strong skill signal.")
    if df_raw.empty: st.info("No data."); return
    s = f.get("snapshot_date")
    cl = snap(df_raw, s); cl = cl[cl.data_quality_flag == "CLEAN"]
    il = cl[["pms_provider","ia_name","service_type"]].drop_duplicates().sort_values(["pms_provider","ia_name"])

    # Pre-filter to Recommended only option
    sig_f = st.toggle("Show Recommended IAs only", value=False, key="nav_rec")
    if sig_f and not signals.empty:
        rec_set = set(zip(signals[signals.signal=="Recommended"].pms_provider,
                          signals[signals.signal=="Recommended"].ia_name,
                          signals[signals.signal=="Recommended"].service_type))
        il = il[il.apply(lambda r:(r.pms_provider,r.ia_name,r.service_type) in rec_set, axis=1)]

    if il.empty: st.info("No IAs match filter."); return
    all_provs_nav = sorted(il.pms_provider.unique().tolist())
    prov_search_nav = st.text_input("🔍 Search provider", "", key="nav_prov_search", placeholder="e.g. Motilal, HDFC, Sundaram...")
    filtered_provs_nav = [p2 for p2 in all_provs_nav if prov_search_nav.lower() in p2.lower()] if prov_search_nav else all_provs_nav
    if not filtered_provs_nav: st.info("No providers match — try a different search."); return
    sel_prov_nav = st.selectbox("Select Provider", filtered_provs_nav, key="nav_prov_sel")
    ia_filtered_nav = il[il.pms_provider == sel_prov_nav]
    ia_opts_nav = [f"{r.ia_name} ({r.service_type})" for r in ia_filtered_nav.itertuples()]
    if not ia_opts_nav: st.info("No strategies for this provider."); return
    sel_ia_nav = st.selectbox("Select Strategy", ia_opts_nav, key="nav_ia_sel")
    sr = ia_filtered_nav.iloc[ia_opts_nav.index(sel_ia_nav)]

    h = pd.read_sql("""SELECT snapshot_date,return_1m FROM raw_performance
        WHERE pms_provider=? AND ia_name=? AND service_type=?
        AND return_1m IS NOT NULL AND (data_quality_flag IS NULL OR data_quality_flag='CLEAN')
        ORDER BY snapshot_date""", C, params=[sr.pms_provider,sr.ia_name,sr.service_type])
    if len(h) < 3: st.info("Not enough history."); return
    h["snapshot_date"] = pd.to_datetime(h["snapshot_date"])
    st0 = h.snapshot_date.min()
    h["NAV"] = 100 * (1 + h.return_1m/100).cumprod()

    def bns(nm):
        q = pd.read_sql("SELECT snapshot_date,return_1m FROM raw_benchmarks WHERE benchmark_name=? AND return_1m IS NOT NULL ORDER BY snapshot_date",
                        C, params=[nm])
        if q.empty: return None
        q["snapshot_date"] = pd.to_datetime(q["snapshot_date"]); q = q[q.snapshot_date >= st0]
        if q.empty: return None
        q["NAV"] = 100 * (1 + q.return_1m/100).cumprod(); return q

    lbl = sr.ia_name[:35]
    # Show signal
    if not signals.empty:
        sr2 = signals[(signals.pms_provider==sr.pms_provider)&(signals.ia_name==sr.ia_name)&(signals.service_type==sr.service_type)]
        if not sr2.empty:
            row = sr2.iloc[0]
            st.markdown(f"**Signal:** {row.signal} &nbsp;|&nbsp; **Score:** {int(row.top_half_months)}/{int(row.months_available)} months top-half")

    fig = go.Figure()
    fig.add_trace(go.Scatter(x=h.snapshot_date, y=h.NAV.round(2), mode="lines", name=lbl,
                             line=dict(color="#2563eb", width=2.5)))
    for nm, cl2, ds in [("Nifty 50 TRI","#f59e0b","dash"),("BSE 500 TRI","#7c3aed","dash"),("Nifty Midcap 150 TRI","#059669","dot")]:
        b = bns(nm)
        if b is not None:
            fig.add_trace(go.Scatter(x=b.snapshot_date, y=b.NAV.round(2), mode="lines",
                                     name=nm, line=dict(color=cl2, dash=ds, width=1.5)))
    fig.add_hline(y=100, line_dash="dot", line_color="#94a3b8", opacity=0.5)
    fig.update_layout(title=f"NAV indexed to 100 — {lbl}", height=400, template="plotly_white",
                      hovermode="x unified", legend=dict(orientation="h", y=-0.25))
    st.plotly_chart(fig, use_container_width=True)

    r = h.return_1m.values; mr, sr3 = np.mean(r), np.std(r, ddof=1)
    sh = (mr-RF)/sr3 if sr3 > 0 else None
    dn = r[r < RF]; dn2 = np.std(dn, ddof=1) if len(dn) > 1 else None
    so = (mr-RF)/dn2 if dn2 and dn2 > 0 else None
    w = np.cumprod(np.insert(1+r/100, 0, 1.0)); pk = np.maximum.accumulate(w)
    md = float(np.min((w[1:]-pk[1:])/pk[1:]*100))
    ni = bns("Nifty 50 TRI"); ir = None
    if ni is not None:
        mg = h.merge(ni[["snapshot_date","return_1m"]].rename(columns={"return_1m":"nf"}), on="snapshot_date", how="inner")
        if len(mg) >= 6:
            act = mg.return_1m.values - mg.nf.values; te = np.std(act, ddof=1)
            ir = float(np.mean(act)/te) if te > 0 else None
    m1,m2,m3,m4,m5 = st.columns(5)
    m1.metric("Sharpe", f"{sh:.2f}" if sh else "—")
    m2.metric("Sortino", f"{so:.2f}" if so else "—")
    m3.metric("Std Dev", f"{sr3:.2f}%")
    m4.metric("Max Drawdown", f"{md:.1f}%")
    m5.metric("Info Ratio", f"{ir:.2f}" if ir else "—", help=">0.5 = strong skill vs benchmark")
    st.caption(f"**{len(r)} months** · {h.snapshot_date.min().strftime('%b %Y')} – {h.snapshot_date.max().strftime('%b %Y')} · RF: 6.5% p.a.")
    nc1, nc2 = st.columns(2)
    with nc2:
        if st.button("📊 Excel — Full history (all periods)", key="xl_nav"):
            xl = excel_single_pms(sr.pms_provider, sr.ia_name, sr.service_type, f.get("snapshot_date"))
            if xl:
                ia_clean = sr.ia_name[:30].replace(' ','_').replace('/','_')
                st.download_button("⬇️ Download Excel", data=xl,
                    file_name=f"PMS_{ia_clean}_nav.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="xl_nav_dl")

# ── TAB 5: QUARTILE ANALYSIS ──────────────────────────────────────────────────
def t_quartile(df, f, db):
    sec("🔢 Rolling Quartile Analysis")
    exp("Two lenses per month: rank by absolute 1M return vs rank by Nifty 50 outperformance — both within strategy peers. "
        "Q1=top 25%. Divergence between the two lenses is itself a signal: "
        "Q1 absolute but Q3 vs Nifty means the market carried the strategy, not the manager.")
    if df.empty: st.info("No data."); return
    s = f.get("snapshot_date")
    cl = snap(df, s); cl = cl[cl.data_quality_flag == "CLEAN"]
    il = cl[["pms_provider","ia_name","service_type","strategy_type"]].drop_duplicates().sort_values(["pms_provider","ia_name"])
    if il.empty: st.info("No data."); return
    all_provs_q = sorted(il.pms_provider.unique().tolist())
    prov_search_q = st.text_input("🔍 Search provider", "", key="q_prov_search", placeholder="e.g. Motilal, HDFC, Sundaram...")
    filtered_provs_q = [p2 for p2 in all_provs_q if prov_search_q.lower() in p2.lower()] if prov_search_q else all_provs_q
    if not filtered_provs_q: st.info("No providers match."); return
    sel_prov_q = st.selectbox("Select Provider", filtered_provs_q, key="q_prov_sel")
    ia_filtered_q = il[il.pms_provider == sel_prov_q]
    ia_opts_q = [f"{r.ia_name} ({r.service_type})" for r in ia_filtered_q.itertuples()]
    if not ia_opts_q: st.info("No strategies for this provider."); return
    sel_ia_q = st.selectbox("Select Strategy", ia_opts_q, key="q_ia_sel")
    sr = ia_filtered_q.iloc[ia_opts_q.index(sel_ia_q)]
    hi = pd.read_sql("""SELECT snapshot_date,return_1m,return_1y FROM raw_performance
        WHERE pms_provider=? AND ia_name=? AND service_type=?
        AND (data_quality_flag IS NULL OR data_quality_flag='CLEAN') ORDER BY snapshot_date""",
        C, params=[sr.pms_provider,sr.ia_name,sr.service_type])
    pe = pd.read_sql("""SELECT snapshot_date,return_1m,return_1y FROM raw_performance
        WHERE strategy_type=? AND (data_quality_flag IS NULL OR data_quality_flag='CLEAN') ORDER BY snapshot_date""",
        C, params=[sr.strategy_type])
    nf = pd.read_sql("SELECT snapshot_date,return_1m AS n1m,return_1y AS n1y FROM raw_benchmarks WHERE benchmark_name='Nifty 50 TRI'", C)
    for d in [hi,pe,nf]: d["snapshot_date"] = pd.to_datetime(d["snapshot_date"])
    pe = pe.merge(nf, on="snapshot_date", how="left")
    pe["a1m"] = pe.return_1m - pe.n1m.fillna(0); pe["a1y"] = pe.return_1y - pe.n1y.fillna(0)
    def qr(v, s2):
        s2 = s2.dropna()
        if len(s2) < 4 or pd.isna(v): return None
        pct = (s2 <= v).sum() / len(s2)
        return 1 if pct >= 0.75 else (2 if pct >= 0.50 else (3 if pct >= 0.25 else 4))
    rows = []
    for _, r in hi.iterrows():
        ps = pe[pe.snapshot_date == r.snapshot_date]
        ns = nf[nf.snapshot_date == r.snapshot_date]
        n1m = float(ns.n1m.iloc[0]) if not ns.empty and pd.notna(ns.n1m.iloc[0]) else None
        n1y = float(ns.n1y.iloc[0]) if not ns.empty and pd.notna(ns.n1y.iloc[0]) else None
        a1m = (r.return_1m - n1m) if n1m and pd.notna(r.return_1m) else None
        a1y = (r.get("return_1y") - n1y) if n1y and pd.notna(r.get("return_1y")) else None
        rows.append({"Month":r.snapshot_date.strftime("%b %Y"),
                     "1M%":r.return_1m,"Q-1M Abs":qr(r.return_1m,ps.return_1m),"Q-1M Nifty":qr(a1m,ps.a1m),
                     "1Y%":r.get("return_1y"),"Q-1Y Abs":qr(r.get("return_1y"),ps.return_1y),"Q-1Y Nifty":qr(a1y,ps.a1y),
                     "#Peers":len(ps)})
        dq = pd.DataFrame(rows)
    dq["_sort"] = pd.to_datetime(dq["Month"], format="%b %Y")
    dq = dq.sort_values("_sort", ascending=False).drop(columns=["_sort"])
    if dq.empty: st.info("Not enough data."); return
    # Side-by-side panels: 1M lens | 1Y lens
    left, right = st.columns(2)
    with left:
        st.markdown("**1M Return Quartiles**")
        st.caption("Rolling 1M return vs category peers — 36 observations")
        d1m = dq[["Month","1M%","Q-1M Abs","Q-1M Nifty","#Peers"]].copy()
        q1m = d1m["Q-1M Abs"].dropna()
        st.dataframe(d1m, use_container_width=True, hide_index=True, height=420)
        st.caption(f"Q1 months (abs): **{(q1m==1).sum()}/{len(q1m)}** · Median Q: **{q1m.median():.1f}**")
    with right:
        st.markdown("**1Y Return Quartiles**")
        st.caption("Rolling 1Y return vs category peers — ~24 observations")
        d1y = dq[["Month","1Y%","Q-1Y Abs","Q-1Y Nifty","#Peers"]].copy()
        q1y = d1y["Q-1Y Abs"].dropna()
        st.dataframe(d1y, use_container_width=True, hide_index=True, height=420)
        st.caption(f"Q1 months (abs): **{(q1y==1).sum()}/{len(q1y)}** · Median Q: **{q1y.median():.1f}**")
    # Check if divergence exists
    q1m_med = dq["Q-1M Abs"].dropna().median() if "Q-1M Abs" in dq.columns else None
    q1y_med = dq["Q-1Y Abs"].dropna().median() if "Q-1Y Abs" in dq.columns else None
    divergence_text = ""
    if q1m_med and q1y_med:
        if abs(q1m_med - q1y_med) >= 1:
            divergence_text = f" · ⚠️ **Divergence detected** — 1M median Q{q1m_med:.0f} vs 1Y median Q{q1y_med:.0f}"
    st.markdown(f"**Strategy: {sr.strategy_type}**{divergence_text}")
    q1 = dq["Q-1M Abs"].dropna()
    dc = dq[["Month","Q-1M Abs","Q-1M Nifty"]].dropna()
    if not dc.empty:
        dm = dc.melt(id_vars="Month", var_name="Lens", value_name="Quartile")
        dm["Month"] = pd.Categorical(dm.Month, categories=dq.Month.tolist()[::-1], ordered=True)
        # Add 1Y lens to chart as well
        dc2 = dq[["Month","Q-1M Abs","Q-1M Nifty","Q-1Y Abs"]].copy()
        dc2 = dc2.melt(id_vars="Month", var_name="Lens", value_name="Quartile").dropna()
        dc2["Month"] = pd.Categorical(dc2["Month"],
                                       categories=dq["Month"].tolist()[::-1], ordered=True)
        dm = dc2
        fig = px.line(dm.sort_values("Month"), x="Month", y="Quartile", color="Lens", markers=True,
                      title="Rolling Quartile Rank (Q1=best)",
                      color_discrete_sequence=["#2563eb","#f59e0b","#16a34a"])
        fig.update_yaxes(autorange="reversed", tickvals=[1,2,3,4], ticktext=["Q1","Q2","Q3","Q4"])
        fig.update_layout(height=340, template="plotly_white", legend=dict(orientation="h", y=-0.25))
        st.plotly_chart(fig, use_container_width=True)
    qc1, qc2 = st.columns(2)
    with qc1:
        csvb(dq, f"quartile_{sr.ia_name[:15].replace(' ','_')}.csv")
    with qc2:
        if st.button("📊 Excel — Full history (all periods)", key="xl_q"):
            xl = excel_single_pms(sr.pms_provider, sr.ia_name, sr.service_type, f.get("snapshot_date"))
            if xl:
                ia_clean = sr.ia_name[:30].replace(' ','_').replace('/','_')
                st.download_button("⬇️ Download Excel", data=xl,
                    file_name=f"PMS_{ia_clean}_quartile.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="xl_q_dl")

# ── TAB 6: AUM TREND ──────────────────────────────────────────────────────────
def t_aum(da, f, dr):
    sec("💰 AUM — Performance vs Investor Flows")
    exp("Two separate series: NAV-driven growth vs net investor flows. Never merged. "
        "Rising AUM from inflows = gaining reputation. Rising AUM from performance = NAV compounding.")
    if da.empty: st.warning("Run calculate_aum_decomposition() first."); return
    mode = st.radio("View", ["📊 Industry","🔎 Fund Deep-Dive"], horizontal=True, key="aum_m")
    strats = f.get("strategies",[]); s = f.get("snapshot_date")
    # Warn if latest month has significantly fewer rows than previous
    aum_counts = da.groupby("snapshot_date").size()
    if len(aum_counts) >= 2:
        latest_count = aum_counts.iloc[-1]
        prev_count   = aum_counts.iloc[-2]
        if latest_count < prev_count * 0.7:
            latest_m = aum_counts.index[-1].strftime("%b %Y")
            st.markdown(f"<div class='wb'>⚠️ <b>{latest_m} data is incomplete</b> — only {latest_count:,} IAs reported vs {prev_count:,} in the previous month. "
                        f"Some providers submit late. The sharp drop in the chart reflects missing data, not actual AUM loss. "
                        f"Switch sidebar to the previous month for complete data.</div>", unsafe_allow_html=True)

    if mode == "📊 Industry":
        df2 = da[da.strategy_type.isin(strats)] if strats else da
        mo = df2.groupby("snapshot_date").agg(organic=("organic_growth","sum"),flows=("net_flows","sum"),total=("aum_current","sum")).reset_index()
        fig = make_subplots(specs=[[{"secondary_y":True}]])
        fig.add_trace(go.Bar(x=mo.snapshot_date,y=mo.organic,name="NAV-driven",marker_color="#16a34a",opacity=0.8),secondary_y=False)
        fig.add_trace(go.Bar(x=mo.snapshot_date,y=mo.flows,name="Net flows",marker_color="#f59e0b",opacity=0.8),secondary_y=False)
        fig.add_trace(go.Scatter(x=mo.snapshot_date,y=mo.total,name="Total AUM",line=dict(color="#1e293b",width=2)),secondary_y=True)
        fig.update_layout(title="Industry AUM", barmode="relative", height=420, template="plotly_white")
        st.plotly_chart(fig, use_container_width=True)
    else:
        cl = snap(dr, s); cl = cl[cl.data_quality_flag=="CLEAN"]
        il = cl[["pms_provider","ia_name","service_type"]].drop_duplicates().sort_values(["pms_provider","ia_name"])
        if il.empty: st.info("No data."); return
        all_provs_aum = sorted(il.pms_provider.unique().tolist())
        prov_search_aum = st.text_input("🔍 Search provider", "", key="aum_prov_search", placeholder="e.g. Motilal, HDFC, Sundaram...")
        filtered_provs_aum = [p2 for p2 in all_provs_aum if prov_search_aum.lower() in p2.lower()] if prov_search_aum else all_provs_aum
        if not filtered_provs_aum: st.info("No providers match."); return
        sel_prov_aum = st.selectbox("Select Provider", filtered_provs_aum, key="aum_prov_sel")
        ia_filtered_aum = il[il.pms_provider == sel_prov_aum]
        ia_opts_aum = [f"{r.ia_name} ({r.service_type})" for r in ia_filtered_aum.itertuples()]
        if not ia_opts_aum: st.info("No strategies for this provider."); return
        sel_ia_aum = st.selectbox("Select Strategy", ia_opts_aum, key="aum_ia_sel")
        sr = ia_filtered_aum.iloc[ia_opts_aum.index(sel_ia_aum)]
        fd = da[(da.pms_provider==sr.pms_provider)&(da.ia_name==sr.ia_name)&(da.service_type==sr.service_type)].sort_values("snapshot_date")
        if fd.empty: st.info("No AUM decomp data for this IA."); return
        lt, od = fd.iloc[-1], fd.iloc[0]
        m1,m2,m3,m4 = st.columns(4)
        m1.metric("Current AUM",f"₹{lt.aum_current:,.0f} Cr"); m2.metric("Start AUM",f"₹{od.aum_current:,.0f} Cr")
        m3.metric("Total Organic",f"₹{fd.organic_growth.sum():,.0f} Cr")
        tf = fd.net_flows.sum(); m4.metric("Net Inflows" if tf>0 else "Net Outflows",f"₹{abs(tf):,.0f} Cr")
        fig = make_subplots(specs=[[{"secondary_y":True}]])
        fig.add_trace(go.Bar(x=fd.snapshot_date,y=fd.organic_growth,name="NAV-driven",marker_color="#16a34a",opacity=0.8),secondary_y=False)
        fig.add_trace(go.Bar(x=fd.snapshot_date,y=fd.net_flows,name="Net flows",marker_color="#f59e0b",opacity=0.8),secondary_y=False)
        fig.add_trace(go.Scatter(x=fd.snapshot_date,y=fd.aum_current,name="AUM",line=dict(color="#1e293b",width=2.5)),secondary_y=True)
        fig.update_layout(title=f"{sr.ia_name} — AUM Decomposition",barmode="relative",height=400,template="plotly_white",legend=dict(orientation="h",y=-0.25))
        st.plotly_chart(fig, use_container_width=True)
        csvb(fd, f"aum_{sr.ia_name[:15].replace(' ','_')}.csv")

# ── TAB 7: FLAGS & ALERTS ─────────────────────────────────────────────────────
def t_flags(df_raw, db, signals, f):
    sec("🚨 Flags & Alerts — Recommended PMSs")
    exp("Monitors all Recommended IAs (auto-signal ≥78% consistency). "
        "Alerts when current snapshot return is below Nifty 50 TRI or category median.")
    if signals.empty: st.info("Not enough data to compute signals."); return
    s = f.get("snapshot_date")
    cl_full = perf_raw()
    cl_full = cl_full[cl_full.data_quality_flag=="CLEAN"]
    latest_per_ia = cl_full.sort_values("snapshot_date").groupby(
        ["pms_provider","ia_name","service_type"]).last().reset_index()
    cl = snap(cl_full, s)
    bs = snap(db, s); nf = bs[bs.benchmark_name.str.contains("Nifty 50", na=False)]
    rec = signals[signals.signal == "Recommended"].sort_values("consistency_score", ascending=False)
    if rec.empty: st.info("No Recommended IAs yet — check Signals tab."); return
    st.markdown(f"**{len(rec)} Recommended IAs** monitored (sorted by consistency score)")
    st.markdown("---")
    alerts = []
    for _, r in rec.iterrows():
        ia = latest_per_ia[(latest_per_ia.pms_provider==r.pms_provider)&(latest_per_ia.ia_name==r.ia_name)&(latest_per_ia.service_type==r.service_type)]
        if ia.empty: continue
        ir2 = ia.iloc[0]; st2 = ir2.get("strategy_type","")
        ia_snap = ir2.get("snapshot_date")
        peers = cl_full[(cl_full.strategy_type==st2)&(cl_full.snapshot_date==ia_snap)]
        for pc, lbl, ic in [("return_1m","1M",False),("return_1y","1Y",True),("return_3y","3Y",True)]:
            if pc not in ir2.index or pd.isna(ir2.get(pc)): continue
            iv = float(ir2[pc])
            if not nf.empty and pc in nf.columns:
                nv = nf[pc].iloc[0]
                if pd.notna(nv) and iv < float(nv):
                    alerts.append({"IA":r.ia_name,"Provider":r.pms_provider,
                                   "Score":f"{int(r.top_half_months)}/{int(r.months_available)}",
                                   "Period":lbl,"IA Return":fp(iv,ic),"vs":"Nifty 50 TRI",
                                   "Bench Return":fp(float(nv),ic),"Gap":fp(iv-float(nv),False)})
            if pc in peers.columns:
                cm = peers[pc].median()
                if pd.notna(cm) and iv < float(cm):
                    alerts.append({"IA":r.ia_name,"Provider":r.pms_provider,
                                   "Score":f"{int(r.top_half_months)}/{int(r.months_available)}",
                                   "Period":lbl,"IA Return":fp(iv,ic),"vs":f"{st2} Median",
                                   "Bench Return":fp(float(cm),ic),"Gap":fp(iv-float(cm),False)})
    if alerts:
        da = pd.DataFrame(alerts)
        st.markdown(f"### ⚠️ {len(da)} underperformance alerts")
        st.dataframe(da, use_container_width=True, hide_index=True)
        csvb(da, "flags_alerts.csv")
    else:
        st.success("✅ All Recommended IAs outperforming benchmarks and category medians.")

    # Rolling trend alert — is a Recommended PMS's consistency score dropping?
    st.markdown("---")
    st.markdown("### 📉 Consistency trend alerts (score dropping over last 6 months)")
    st.caption("Compares each Recommended IA's consistency score in the last 6 months vs the 6 months before that. A falling score is an early warning before returns deteriorate.")
    trend_alerts = []
    full_perf = perf_raw()
    all_dates = sorted(full_perf["snapshot_date"].unique(), reverse=True)
    recent_6 = all_dates[:6]; prev_6 = all_dates[6:12]
    nifty_m2 = (bench_data()[bench_data().benchmark_name.str.contains("Nifty 50",na=False)]
                [["snapshot_date","return_1m"]].rename(columns={"return_1m":"nifty_1m"}))
    for _, r in rec.iterrows():
        for window, label in [(recent_6,"Last 6M"),(prev_6,"Prev 6M")]:
            pass
        ia_hist = full_perf[(full_perf.pms_provider==r.pms_provider)&
                            (full_perf.ia_name==r.ia_name)&
                            (full_perf.service_type==r.service_type)&
                            (full_perf.data_quality_flag=="CLEAN")].copy()
        ia_hist = ia_hist.merge(nifty_m2, on="snapshot_date", how="left")
        ia_hist["alpha_1m"] = ia_hist["return_1m"] - ia_hist["nifty_1m"].fillna(0)
        strat2 = signals[signals.ia_name==r.ia_name]["strategy_type"].iloc[0] if not signals[signals.ia_name==r.ia_name].empty else ""
        def score_window(dates):
            sub = ia_hist[ia_hist.snapshot_date.isin(dates)]
            if len(sub) < 3: return None
            hits = 0
            for _, row2 in sub.iterrows():
                peers = full_perf[(full_perf.snapshot_date==row2.snapshot_date)&
                                  (full_perf.strategy_type==strat2)&
                                  (full_perf.data_quality_flag=="CLEAN")]
                peers = peers.merge(nifty_m2, on="snapshot_date", how="left")
                peers["alpha_1m"] = peers["return_1m"] - peers["nifty_1m"].fillna(0)
                abs_pct = (peers["return_1m"].dropna()<=row2.return_1m).sum()/max(len(peers["return_1m"].dropna()),1) if pd.notna(row2.get("return_1m")) else 0
                alp_pct = (peers["alpha_1m"].dropna()<=row2.alpha_1m).sum()/max(len(peers["alpha_1m"].dropna()),1) if pd.notna(row2.get("alpha_1m")) else 0
                if abs_pct >= 0.50 and alp_pct >= 0.50: hits += 1
            return round(hits/len(sub)*100, 1)
        sc_recent = score_window(recent_6)
        sc_prev   = score_window(prev_6)
        if sc_recent is not None and sc_prev is not None:
            drop = sc_prev - sc_recent
            if drop >= 20:
                trend_alerts.append({
                    "IA": r.ia_name, "Provider": r.pms_provider,
                    "Prev 6M Score": f"{sc_prev}%", "Last 6M Score": f"{sc_recent}%",
                    "Drop": f"-{drop:.0f}pp",
                    "Status": "🔴 Sharp drop" if drop >= 35 else "🟡 Declining"
                })
    if trend_alerts:
        st.dataframe(pd.DataFrame(trend_alerts), use_container_width=True, hide_index=True)
    else:
        st.success("✅ No Recommended IAs showing significant consistency decline.")
    st.markdown("---")
    st.markdown("### All Recommended IAs — ranked by consistency score")
    rows = []
    for _, r in rec.iterrows():
        ia = latest_per_ia[(latest_per_ia.pms_provider==r.pms_provider)&(latest_per_ia.ia_name==r.ia_name)&(latest_per_ia.service_type==r.service_type)]
        if ia.empty: continue
        row = ia.iloc[0]
        rows.append({"Score":f"{int(r.top_half_months)}/{int(r.months_available)}",
                     "Score%":f"{r.consistency_score*100:.0f}%",
                     "Provider":r.pms_provider,"IA":r.ia_name,"Strategy":r.strategy_type,
                     "AUM(₹Cr)":round(row.aum_cr,1) if pd.notna(row.get("aum_cr")) else None,
                     "1M%":row.get("return_1m"),"1Y%":row.get("return_1y"),"3Y%":row.get("return_3y")})
    if rows:
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
        csvb(pd.DataFrame(rows), "recommended_ias.csv")

# ── TAB 8: IA COMPARE ─────────────────────────────────────────────────────────
def t_compare(dr, dri, f, db, signals):
    sec("🔀 IA Compare")
    exp("Compare up to 5 IAs. Auto-signal and consistency score shown for each. "
        "Use this to compare two Recommended IAs — higher score = more consistent over 18 months.")
    cl = dr[dr.data_quality_flag=="CLEAN"]; lt = cl.snapshot_date.max(); dl = cl[cl.snapshot_date==lt].copy()
    dl["lbl"] = dl.pms_provider + " · " + dl.ia_name
    sel = st.multiselect("Select IAs (max 5)", sorted(dl.lbl.unique()), max_selections=5, key="cmp_sel")
    if not sel: st.info("Select at least one IA."); return
    ds = dl[dl.lbl.isin(sel)].copy(); ians = ds.ia_name.tolist()
    RC = ["return_1m","return_3m","return_6m","return_1y","return_3y","return_5y","return_si"]
    RL = ["1M","3M","6M","1Y","3Y","5Y","SI"]; RC2 = {"1Y","3Y","5Y","SI"}
    rows = []
    for _, r in ds.iterrows():
        sig_row = signals[(signals.pms_provider==r.pms_provider)&(signals.ia_name==r.ia_name)&(signals.service_type==r.service_type)] if not signals.empty else pd.DataFrame()
        sig = sig_row.iloc[0].signal if not sig_row.empty else "—"
        score = f"{int(sig_row.iloc[0].top_half_months)}/{int(sig_row.iloc[0].months_available)}" if not sig_row.empty else "—"
        row = {"Signal":sig,"Score":score,"Provider":r.pms_provider,"IA":r.ia_name,
               "Strategy":r.get("strategy_type","—"),"AUM(₹Cr)":f"{r.aum_cr:,.0f}" if pd.notna(r.get("aum_cr")) else "—"}
        for c, l in zip(RC, RL): v = r.get(c); row[l] = fp(v, l in RC2) if pd.notna(v) else "—"
        rows.append(row)
    bs = snap(db, f.get("snapshot_date")); nb = bs[bs.benchmark_name.str.contains("Nifty 50", na=False)]
    if not nb.empty:
        row = {"Signal":"—","Score":"—","Provider":"Benchmark","IA":"Nifty 50 TRI","Strategy":"—","AUM(₹Cr)":"—"}
        for c, l in zip(RC, RL): v = nb[c].iloc[0] if c in nb.columns else None; row[l] = fp(v, l in RC2) if pd.notna(v) else "—"
        rows.append(row)
    st.markdown("### Side-by-side snapshot")
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
    bd = []
    for _, r in ds.iterrows():
        lb2 = r.pms_provider + " · " + r.ia_name
        for c, p2 in zip(RC, RL):
            v = r.get(c)
            if pd.notna(v): bd.append({"IA":lb2,"Period":p2,"Return(%)":v})
    if bd:
        df3 = pd.DataFrame(bd); df3["Period"] = pd.Categorical(df3.Period, categories=RL, ordered=True)
        fig = px.bar(df3.sort_values("Period"), x="Period", y="Return(%)", color="IA", barmode="group",
                     color_discrete_sequence=px.colors.qualitative.Set2)
        fig.update_layout(height=360, template="plotly_white", legend=dict(orientation="h",y=-0.3))
        st.plotly_chart(fig, use_container_width=True)
    st.markdown("### 1M Return History")
    dh = cl[cl.ia_name.isin(ians)].copy(); dh["snapshot_date"] = pd.to_datetime(dh["snapshot_date"])
    fig2 = go.Figure()
    for ia in ians:
        d = dh[dh.ia_name==ia].sort_values("snapshot_date")
        if not d.empty and "return_1m" in d.columns:
            fig2.add_trace(go.Scatter(x=d.snapshot_date,y=d.return_1m,mode="lines+markers",
                                      name=f"{d.pms_provider.iloc[0]} · {ia}",line=dict(width=2)))
    fig2.update_layout(height=400, template="plotly_white", legend=dict(orientation="h",y=-0.3))
    st.plotly_chart(fig2, use_container_width=True)

    # Consistency score comparison chart
    if not signals.empty:
        st.markdown("### Consistency score comparison (last 18 months)")
        sc_rows = []
        for _, r in ds.iterrows():
            sr2 = signals[(signals.pms_provider==r.pms_provider)&(signals.ia_name==r.ia_name)&(signals.service_type==r.service_type)]
            if not sr2.empty:
                sc_rows.append({"IA":r.ia_name,"Score %":sr2.iloc[0].consistency_score*100,
                                 "Signal":sr2.iloc[0].signal})
        if sc_rows:
            df_sc = pd.DataFrame(sc_rows).sort_values("Score %", ascending=True)
            fig3 = px.bar(df_sc, y="IA", x="Score %", color="Signal", orientation="h",
                          color_discrete_map={"Recommended":"#16a34a","Hold":"#d97706","Not Recommended":"#dc2626"},
                          title="% months in top-half vs peers (both lenses)")
            fig3.add_vline(x=78, line_dash="dash", line_color="#16a34a", annotation_text="Rec threshold")
            fig3.add_vline(x=50, line_dash="dash", line_color="#d97706", annotation_text="Hold threshold")
            fig3.update_layout(height=300, template="plotly_white")
            st.plotly_chart(fig3, use_container_width=True)
    csvb(pd.DataFrame(rows), "ia_compare.csv", "⬇️ CSV")

# ── TAB 9: TIME-SERIES ────────────────────────────────────────────────────────
def t_ts(df, f, db):
    sec("📈 Time-Series — Monthly Return Evolution")
    exp("Track up to 5 IAs across all 37 months. Overlay Nifty 50 TRI.")
    if df.empty: st.info("No data."); return
    p, pl = f["period"], f["period_label"]
    c1,c2,c3 = st.columns(3)
    with c1: strats = st.multiselect("Strategy", df.strategy_type.unique().tolist(), default=["Equity"], key="ts_s")
    with c2:
        provs = sorted(df[df.strategy_type.isin(strats)].pms_provider.unique()) if strats else []
        sp = st.multiselect("Provider", provs, default=provs[:2] if provs else [], key="ts_p")
    with c3:
        ias = sorted(df[df.pms_provider.isin(sp)].ia_name.unique()) if sp else []
        si = st.multiselect(f"IA ({len(ias)})", ias, default=ias[:3] if ias else [], max_selections=5, key="ts_i")
    sb = st.toggle("Overlay benchmarks (Nifty 50, BSE 500, Nifty Midcap 150)", value=True, key="ts_b")
    if not si: st.info("Select at least one IA."); return
    if p not in df.columns: st.warning(f"{pl} not available."); return
    dt = df[df.ia_name.isin(si)].dropna(subset=[p]).sort_values("snapshot_date")
    fig = go.Figure()
    for ia in si:
        sub = dt[dt.ia_name==ia]
        if sub.empty: continue
        fig.add_trace(go.Scatter(x=sub.snapshot_date,y=sub[p],mode="lines+markers",
                                 name=f"{sub.pms_provider.iloc[0]} · {ia}",
                                 hovertemplate=f"<b>{ia}</b><br>%{{x|%b %Y}}: %{{y:.2f}}%<extra></extra>",
                                 line=dict(width=2),marker=dict(size=4)))
    if sb and not db.empty:
        bench_ts_styles = {
            "Nifty 50 TRI":         ("#f59e0b","dash"),
            "BSE 500 TRI":          ("#7c3aed","dot"),
            "Nifty Midcap 150 TRI": ("#059669","dashdot"),
        }
        for bname, (bcolor, bdash) in bench_ts_styles.items():
            bdf = db[db.benchmark_name==bname]
            if not bdf.empty and p in bdf.columns:
                bdf2 = bdf.dropna(subset=[p])
                fig.add_trace(go.Scatter(
                    x=bdf2.snapshot_date, y=bdf2[p],
                    mode="lines", name=bname,
                    line=dict(color=bcolor, dash=bdash, width=1.8)))
    fig.add_hline(y=0, line_dash="dot", line_color="#94a3b8", opacity=0.5)
    fig.update_layout(title=f"{pl} Return — Selected IAs (trailing, point-in-time)",
                      height=420, template="plotly_white",
                      hovermode="x unified", legend=dict(orientation="h",y=-0.3))
    st.plotly_chart(fig, use_container_width=True)

    # ── COMPOUNDED NAV CHART ───────────────────────────────────────────────
    st.markdown("### 📈 Compounded wealth — ₹100 invested at start")
    st.caption("Shows how ₹100 invested on the first available date would have grown. "
               "This is the honest picture of total wealth created, independent of recent momentum. "
               "Each line = one PMS, compounded month by month from 1M TWRR returns.")

    nav_fig = go.Figure()
    nifty_nav = pd.read_sql("""
        SELECT snapshot_date, return_1m FROM raw_benchmarks
        WHERE benchmark_name = 'Nifty 50 TRI' AND return_1m IS NOT NULL
        ORDER BY snapshot_date
    """, C)
    nifty_nav["snapshot_date"] = pd.to_datetime(nifty_nav["snapshot_date"])

    # Find common start date across all selected IAs
    start_dates = []
    ia_histories = {}
    for ia in si:
        sub = df[df.ia_name == ia].dropna(subset=["return_1m"]).sort_values("snapshot_date")
        if not sub.empty:
            start_dates.append(sub.snapshot_date.min())
            ia_histories[ia] = sub

    if start_dates:
        common_start = max(start_dates)  # start where all IAs have data

        colors_nav = ["#2563eb","#dc2626","#16a34a","#d97706","#7c3aed"]
        for idx2, ia in enumerate(si):
            if ia not in ia_histories: continue
            h = ia_histories[ia]
            h = h[h.snapshot_date >= common_start].copy()
            if len(h) < 2: continue
            h["NAV"] = 100 * (1 + h["return_1m"] / 100).cumprod()
            prov = h.pms_provider.iloc[0]
            nav_fig.add_trace(go.Scatter(
                x=h.snapshot_date, y=h.NAV.round(2),
                mode="lines+markers",
                name=f"{prov[:20]} · {ia[:25]}",
                line=dict(color=colors_nav[idx2 % len(colors_nav)], width=2.5),
                marker=dict(size=4),
                hovertemplate=f"<b>{ia[:30]}</b><br>%{{x|%b %Y}}: ₹%{{y:.1f}}<extra></extra>"
            ))

        # Add Nifty 50 TRI
        if sb:
            nifty_sub = nifty_nav[nifty_nav.snapshot_date >= common_start].copy()
            if not nifty_sub.empty:
                nifty_sub["NAV"] = 100 * (1 + nifty_sub["return_1m"] / 100).cumprod()
                nav_fig.add_trace(go.Scatter(
                    x=nifty_sub.snapshot_date, y=nifty_sub.NAV.round(2),
                    mode="lines", name="Nifty 50 TRI",
                    line=dict(color="#f59e0b", dash="dash", width=2),
                    hovertemplate="<b>Nifty 50 TRI</b><br>%{x|%b %Y}: ₹%{y:.1f}<extra></extra>"
                ))

        nav_fig.add_hline(y=100, line_dash="dot", line_color="#94a3b8",
                          opacity=0.5, annotation_text="₹100 invested")
        nav_fig.update_layout(
            title=f"₹100 invested at {common_start.strftime('%b %Y')} — compounded monthly",
            xaxis_title="Date", yaxis_title="Portfolio value (₹)",
            height=420, template="plotly_white",
            hovermode="x unified",
            legend=dict(orientation="h", y=-0.3)
        )
        st.plotly_chart(nav_fig, use_container_width=True)

        # Final values summary
        st.markdown("**Final portfolio value (₹100 invested at start):**")
        final_cols = st.columns(len(si) + (1 if sb else 0))
        for idx2, ia in enumerate(si):
            if ia not in ia_histories: continue
            h = ia_histories[ia][ia_histories[ia].snapshot_date >= common_start].copy()
            if len(h) < 2: continue
            h["NAV"] = 100 * (1 + h["return_1m"] / 100).cumprod()
            final_val = h["NAV"].iloc[-1]
            total_ret = final_val - 100
            color = "#16a34a" if total_ret > 0 else "#dc2626"
            final_cols[idx2].markdown(
                f"<div style='text-align:center'>"
                f"<div style='font-size:11px;color:#64748b'>{ia[:25]}</div>"
                f"<div style='font-size:22px;font-weight:700;color:{color}'>₹{final_val:.1f}</div>"
                f"<div style='font-size:12px;color:{color}'>{total_ret:+.1f}% total</div>"
                f"</div>", unsafe_allow_html=True
            )
        if sb:
            bench_summary = pd.read_sql("""
                SELECT snapshot_date, benchmark_name, return_1m FROM raw_benchmarks
                WHERE benchmark_name IN ('Nifty 50 TRI','BSE 500 TRI','Nifty Midcap 150 TRI')
                AND return_1m IS NOT NULL ORDER BY snapshot_date
            """, C)
            bench_summary["snapshot_date"] = pd.to_datetime(bench_summary["snapshot_date"])
            bench_colors_sum = {"Nifty 50 TRI":"#f59e0b","BSE 500 TRI":"#7c3aed","Nifty Midcap 150 TRI":"#059669"}
            extra_cols = st.columns(3)
            for bidx, bname in enumerate(["Nifty 50 TRI","BSE 500 TRI","Nifty Midcap 150 TRI"]):
                bs3 = bench_summary[(bench_summary.benchmark_name==bname) &
                                    (bench_summary.snapshot_date >= common_start)].copy()
                if bs3.empty: continue
                bs3["NAV"] = 100 * (1 + bs3["return_1m"] / 100).cumprod()
                bv = bs3["NAV"].iloc[-1]; bc = bench_colors_sum[bname]
                extra_cols[bidx].markdown(
                    f"<div style='text-align:center'>"
                    f"<div style='font-size:11px;color:#64748b'>{bname}</div>"
                    f"<div style='font-size:22px;font-weight:700;color:{bc}'>₹{bv:.1f}</div>"
                    f"<div style='font-size:12px;color:{bc}'>{bv-100:+.1f}% total</div>"
                    f"</div>", unsafe_allow_html=True
                )
        st.caption(f"Common start date: {common_start.strftime('%B %Y')} · "
                   "All returns computed from monthly 1M TWRR, compounded. "
                   "Start date = latest first-available month across all selected IAs.")
    if not dt.empty:
        tc1, tc2 = st.columns(2)
        with tc1:
            csvb(dt[["snapshot_date","pms_provider","ia_name",p,"aum_cr"]], f"timeseries_{pl}.csv", f"⬇️ CSV — {pl}")
        with tc2:
            if len(si) == 1 and st.button("📊 Excel — Full history", key="xl_ts"):
                ia_ts = dt[dt.ia_name==si[0]].iloc[0]
                xl = excel_single_pms(ia_ts.pms_provider, ia_ts.ia_name, ia_ts.service_type, f.get("snapshot_date"))
                if xl:
                    ia_clean = si[0][:30].replace(' ','_').replace('/','_')
                    st.download_button("⬇️ Download Excel", data=xl,
                        file_name=f"PMS_{ia_clean}_timeseries.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="xl_ts_dl")

# ── TAB 10: RISK METRICS ──────────────────────────────────────────────────────
def t_risk(dr, dp, f):
    sec("⚖️ Risk-Adjusted Performance")
    exp("Sharpe=return/vol. Sortino=return/downside vol. Alpha=excess return vs Nifty 50. Info Ratio=alpha÷tracking error.")
    if dr.empty: st.warning("Run calculate_risk_metrics() first."); return
    strats = f.get("strategies",[]); df2 = dr[dr.strategy_type.isin(strats)] if strats else dr
    m1,m2,m3,m4 = st.columns(4)
    m1.metric("Avg Sharpe",f"{df2.sharpe_ratio.mean():.2f}"); m2.metric("Avg Sortino",f"{df2.sortino_ratio.mean():.2f}")
    m3.metric("Avg Max DD",f"{df2.max_drawdown.mean():.2f}%")
    pct = (df2.alpha_vs_nifty50 > 0).mean()*100 if df2.alpha_vs_nifty50.notna().any() else 0
    m4.metric("% +ve Alpha",f"{pct:.1f}%")
    c1,c2 = st.columns(2)
    with c1:
        st.markdown("#### Is higher return worth the risk?")
        st.caption("Each bar = one strategy type. Green = most funds have good risk-adjusted returns. Red = most funds are taking more risk than the return justifies.")
        if "return_1y" in dp.columns:
            dj = df2.merge(dp[["pms_provider","ia_name","service_type","return_1y"]].dropna(),on=["pms_provider","ia_name","service_type"],how="inner")
            if not dj.empty:
                risk_summary = dj.groupby("strategy_type").apply(lambda g: pd.Series({
                    "Good risk/return (Sharpe>0)": (g.sharpe_ratio>0).sum(),
                    "Poor risk/return (Sharpe≤0)": (g.sharpe_ratio<=0).sum(),
                    "Total": len(g)
                })).reset_index()
                risk_summary["% Good"] = (risk_summary["Good risk/return (Sharpe>0)"] / risk_summary["Total"] * 100).round(0)
                fig = px.bar(risk_summary, x="strategy_type", y=["Good risk/return (Sharpe>0)","Poor risk/return (Sharpe≤0)"],
                             barmode="stack",
                             color_discrete_map={"Good risk/return (Sharpe>0)":"#16a34a","Poor risk/return (Sharpe≤0)":"#dc2626"},
                             title="How many funds have good risk-adjusted returns?",
                             labels={"strategy_type":"Strategy","value":"Number of funds"})
                fig.update_layout(height=380,template="plotly_white",legend=dict(orientation="h",y=-0.3))
                st.plotly_chart(fig,use_container_width=True)
                for _, r in risk_summary.iterrows():
                    pct = int(r["% Good"])
                    color = "#16a34a" if pct>=60 else ("#d97706" if pct>=40 else "#dc2626")
                    st.markdown(f"<span style='color:{color}'>{'✅' if pct>=60 else ('⚠️' if pct>=40 else '❌')} <b>{r.strategy_type}:</b> {pct}% of funds are generating returns worth the risk taken.</span>", unsafe_allow_html=True)
    with c2:
        st.markdown("#### What's the worst loss clients could have seen?")
        st.caption("Shows the average worst loss from peak for each strategy type. The smaller the number, the better — it means clients didn't have to stomach big drops.")
        avg_dd = df2.groupby("strategy_type")["max_drawdown"].mean().reset_index()
        avg_dd.columns = ["Strategy","Avg Worst Loss (%)"]
        avg_dd["Avg Worst Loss (%)"] = avg_dd["Avg Worst Loss (%)"].round(1)
        avg_dd["Label"] = avg_dd["Avg Worst Loss (%)"].apply(lambda x: f"{x:.1f}%")
        avg_dd["Color"] = avg_dd["Avg Worst Loss (%)"].apply(
            lambda x: "#16a34a" if x>-10 else ("#d97706" if x>-20 else "#dc2626"))
        fig2 = px.bar(avg_dd, x="Strategy", y="Avg Worst Loss (%)",
                      color="Strategy", color_discrete_map={r.Strategy:r.Color for _,r in avg_dd.iterrows()},
                      text="Label", title="Average worst loss from peak — by strategy type")
        fig2.update_traces(textposition="outside")
        fig2.update_layout(height=380,template="plotly_white",showlegend=False,
                           yaxis_title="Avg worst loss (%)")
        st.plotly_chart(fig2,use_container_width=True)
        for _, r in avg_dd.iterrows():
            v = r["Avg Worst Loss (%)"]
            msg = "minimal losses" if v>-10 else ("moderate losses" if v>-20 else "significant losses")
            color = "#16a34a" if v>-10 else ("#d97706" if v>-20 else "#dc2626")
            st.markdown(f"<span style='color:{color}'>{'✅' if v>-10 else ('⚠️' if v>-20 else '❌')} <b>{r.Strategy}:</b> Clients here experienced {msg} on average (worst drop: {v:.1f}%).</span>", unsafe_allow_html=True)
    cols = [c for c in ["pms_provider","ia_name","strategy_type","sharpe_ratio","sortino_ratio","max_drawdown","alpha_vs_nifty50","beta","info_ratio","std_deviation","months_of_data"] if c in df2.columns]
    disp = df2[cols].copy()
    # Format before rename
    for nc in ["sortino_ratio","max_drawdown","alpha_vs_nifty50","beta","info_ratio"]:
        if nc in disp.columns:
            disp[nc] = disp[nc].apply(lambda x: "—" if pd.isna(x) else round(float(x),3))
    disp.columns = ["Provider","IA","Strategy","Sharpe","Sortino","MaxDD%","Alpha","Beta","IR","StdDev","Months"][:len(cols)]
    st.dataframe(disp.sort_values("Sharpe",ascending=False).round(3),use_container_width=True,height=380,hide_index=True)
    csvb(disp,"risk_metrics.csv","⬇️ CSV")

# ── TAB 11: TURNOVER ──────────────────────────────────────────────────────────
def t_turn(dt, dp, f):
    sec("🔄 Turnover Analysis")
    exp("100% = full portfolio replaced annually. High turnover = higher transaction costs.")
    if dt.empty: st.warning("Run turnover scraper first."); return
    p, pl = f["period"], f["period_label"]; strats = f.get("strategies",[]); s = f.get("snapshot_date")
    df2 = snap(dt, s)
    if strats: df2 = df2[df2.strategy_type.isin(strats)]
    if df2.empty: st.info("No data."); return
    m1,m2,m3,m4 = st.columns(4)
    m1.metric("IAs",f"{len(df2):,}"); m2.metric("Avg",f"{df2.turnover_ratio.mean():.1f}%")
    m3.metric("Median",f"{df2.turnover_ratio.median():.1f}%"); m4.metric(">200%",f"{(df2.turnover_ratio>200).sum():,}")
    dn = df2[df2.turnover_ratio<=300]
    c1,c2 = st.columns(2)
    with c1:
        st.markdown("#### How actively are funds trading?")
        st.caption("Low = trades rarely (cheaper, less disruption). High = trades a lot (more costs). This shows what portion of funds in each category trade at each level.")
        def turnover_bucket(x):
            if x < 30: return "🟢 Low (<30%)"
            elif x < 100: return "🟡 Medium (30-100%)"
            else: return "🔴 High (>100%)"
        dn2 = dn.copy(); dn2["Trading Activity"] = dn2["turnover_ratio"].apply(turnover_bucket)
        bucket_summary = dn2.groupby(["strategy_type","Trading Activity"]).size().reset_index(name="Funds")
        fig = px.bar(bucket_summary, x="strategy_type", y="Funds", color="Trading Activity",
                     barmode="stack",
                     color_discrete_map={"🟢 Low (<30%)":"#16a34a","🟡 Medium (30-100%)":"#d97706","🔴 High (>100%)":"#dc2626"},
                     title="How actively are funds trading — by strategy?",
                     labels={"strategy_type":"Strategy"})
        fig.update_layout(height=360,template="plotly_white",legend=dict(orientation="h",y=-0.3))
        st.plotly_chart(fig,use_container_width=True)
        for strat in dn2.strategy_type.unique():
            sub = dn2[dn2.strategy_type==strat]
            high_pct = (sub["turnover_ratio"]>100).mean()*100
            color = "#16a34a" if high_pct<20 else ("#d97706" if high_pct<50 else "#dc2626")
            st.markdown(f"<span style='color:{color}'><b>{strat}:</b> {high_pct:.0f}% of funds trade heavily — {'manageable' if high_pct<20 else ('worth monitoring' if high_pct<50 else 'high cost drag likely')}.</span>", unsafe_allow_html=True)
    with c2:
        t20 = df2.nlargest(20,"turnover_ratio")[["pms_provider","ia_name","strategy_type","turnover_ratio"]].rename(columns={"pms_provider":"Provider","ia_name":"IA","strategy_type":"Strategy","turnover_ratio":"Turnover %"})
        st.markdown("**Top 20 Highest Turnover**"); st.dataframe(t20.round(1),use_container_width=True,hide_index=True)
    if not dp.empty and p in dp.columns:
        ss = s if s else df2.snapshot_date.max().strftime("%Y-%m-%d")
        dp2 = dp[dp.snapshot_date.dt.strftime("%Y-%m-%d")==ss]
        dm = dn.merge(dp2[["pms_provider","ia_name","service_type",p]].dropna(),on=["pms_provider","ia_name","service_type"],how="inner")
        if not dm.empty:
            st.markdown(f"### Does trading more actually help returns?")
            st.caption("Each bar compares returns of low-turnover vs high-turnover funds within each strategy. If the green bar is taller, funds that trade less are doing better.")
            dm["Trading Activity"] = dm["turnover_ratio"].apply(
                lambda x: "Low turnover (<30%)" if x<30 else ("Medium (30-100%)" if x<100 else "High (>100%)"))
            verdict_data = dm.groupby(["strategy_type","Trading Activity"])[p].median().reset_index()
            verdict_data.columns = ["Strategy","Trading Activity",f"Median {pl} Return (%)"]
            fig2 = px.bar(verdict_data, x="Strategy", y=f"Median {pl} Return (%)",
                          color="Trading Activity", barmode="group",
                          color_discrete_map={"Low turnover (<30%)":"#16a34a","Medium (30-100%)":"#d97706","High (>100%)":"#dc2626"},
                          title=f"Do funds that trade more earn more? — {pl} median returns by trading activity")
            fig2.add_hline(y=0,line_dash="dash",line_color="#94a3b8",opacity=0.4)
            fig2.update_layout(height=380,template="plotly_white",legend=dict(orientation="h",y=-0.3))
            st.plotly_chart(fig2,use_container_width=True)
            # Plain English verdict per strategy
            for strat in dm.strategy_type.unique():
                sub = dm[dm.strategy_type==strat]
                low_ret = sub[sub.turnover_ratio<30][p].median()
                high_ret = sub[sub.turnover_ratio>100][p].median()
                if pd.notna(low_ret) and pd.notna(high_ret):
                    if low_ret > high_ret:
                        st.markdown(f"✅ **{strat}:** Funds that trade less returned {low_ret:.1f}% vs {high_ret:.1f}% for high-turnover funds — trading less worked better.")
                    elif high_ret > low_ret:
                        st.markdown(f"⚠️ **{strat}:** High-turnover funds returned {high_ret:.1f}% vs {low_ret:.1f}% — but higher costs may offset this over time.")
                    else:
                        st.markdown(f"➡️ **{strat}:** No clear difference between low and high turnover funds on returns.")
    csvb(df2,"turnover.csv")

# ── TAB 12: PROVIDER COMPARE ──────────────────────────────────────────────────
def t_provider(dp, dr, f, db):
    sec("🏢 Provider Compare")
    exp("Firm-level metrics aggregated across all CLEAN IAs.")
    cl = dp[dp.data_quality_flag=="CLEAN"].copy(); p, pl = f["period"], f["period_label"]
    strats = f.get("strategies",[]); s = f.get("snapshot_date")
    ds = snap(cl, s)
    if strats: ds = ds[ds.strategy_type.isin(strats)]
    if ds.empty or p not in ds.columns: st.info("No data."); return
    agg = ds.groupby("pms_provider").agg(n=("ia_name","nunique"),aum=("aum_cr","sum"),
        med=(p,"median"),best=(p,"max"),worst=(p,"min"),
        pct=(p,lambda x:(x>0).mean()*100)).round(2).reset_index()
    if not dr.empty:
        ra = dr.sort_values("snapshot_date").groupby("pms_provider").last().reset_index()
        mc = [c for c in ["pms_provider","sharpe_ratio","max_drawdown","alpha_vs_nifty50","info_ratio"] if c in ra.columns]
        agg = agg.merge(ra[mc].round(3), on="pms_provider", how="left")
    rn = {"pms_provider":"Provider","n":"#IAs","aum":"Total AUM(₹Cr)","med":f"Median({pl})",
          "best":f"Best({pl})","worst":f"Worst({pl})","pct":"% Positive",
          "sharpe_ratio":"Sharpe","max_drawdown":"MaxDD%","alpha_vs_nifty50":"Alpha","info_ratio":"IR"}
    lb = agg.rename(columns=rn).sort_values(f"Median({pl})",ascending=False).reset_index(drop=True)
    lb.insert(0,"Rank",range(1,len(lb)+1))
    dcols = ["Rank"] + [v for v in rn.values() if v in lb.columns]
    st.dataframe(lb[dcols],use_container_width=True,height=480,hide_index=True)
    csvb(lb,f"providers_{pl}.csv",f"⬇️ CSV — {pl}")
    st.markdown("---"); st.markdown("### Head-to-head")
    sp2 = st.multiselect("Select providers (max 6)",sorted(agg.pms_provider.unique()),max_selections=6,key="ph2h")
    if not sp2: st.info("Select providers."); return
    di = ds[ds.pms_provider.isin(sp2)].dropna(subset=[p])
    if di.groupby('pms_provider').size().max() < 2:
        st.info('Selected providers have only 1 IA each — box plot needs 2+. Showing bar chart instead.')
        fig_bar = px.bar(di, x='pms_provider', y=p, color='pms_provider', title=f'{pl} Return per Provider')
        fig_bar.update_layout(height=350, template='plotly_white', showlegend=False)
        st.plotly_chart(fig_bar, use_container_width=True)
    elif not di.empty:
        bv = None; bs = snap(db,s); nb = bs[bs.benchmark_name.str.contains("Nifty 50",na=False)]
        if not nb.empty and p in nb.columns and pd.notna(nb[p].iloc[0]): bv = float(nb[p].iloc[0])
        fig = px.box(di,x="pms_provider",y=p,color="pms_provider",points="outliers",
                     hover_data={"ia_name":True},title=f"{pl} Return — All IAs per Provider")
        fig.add_hline(y=0,line_dash="dash",line_color="#6b7280",opacity=0.4)
        if bv: fig.add_hline(y=bv,line_dash="dot",line_color="#f59e0b",opacity=0.8,annotation_text=f"Nifty50:{bv:.2f}%",annotation_font_color="#f59e0b")
        fig.update_layout(height=400,template="plotly_white",showlegend=False,xaxis_tickangle=-20)
        st.plotly_chart(fig,use_container_width=True)
    da2 = cl[cl.pms_provider.isin(sp2)]
    if "aum_cr" in da2.columns and not da2.empty:
        aa = da2.groupby(["snapshot_date","pms_provider"]).aum_cr.sum().reset_index()
        fig2 = px.line(aa,x="snapshot_date",y="aum_cr",color="pms_provider",title="AUM per Provider — All Months")
        fig2.update_layout(height=340,template="plotly_white",hovermode="x unified",legend=dict(orientation="h",y=-0.3))
        st.plotly_chart(fig2,use_container_width=True)

# ── TAB 13: DATA VALIDATION ──────────────────────────────────────────────────
def t_validation(df_perf, df_bench, df_turn):
    from datetime import datetime
    sec("🔍 Data Validation — Scrape Health Check")
    exp("Run this after every monthly scrape to confirm data looks right before sharing with the team. "
        "Green = all good. Yellow = worth checking. Red = something is wrong.")

    st.markdown("### 📦 Coverage")
    c1,c2,c3,c4,c5 = st.columns(5)
    c1.metric("Total rows", f"{len(df_perf):,}", help="Expected ~52,000+")
    c2.metric("Providers", f"{df_perf.pms_provider.nunique():,}", help="Expected ~390-410")
    c3.metric("IAs", f"{df_perf.ia_name.nunique():,}", help="Expected ~2,000+")
    months = df_perf.snapshot_date.nunique()
    c4.metric("Months", f"{months}", help="Expected 37 (Apr 2023 – Apr 2026)")
    c5.metric("Latest snapshot", df_perf.snapshot_date.max().strftime("%b %Y"))

    # Month completeness check
    st.markdown("### 📅 Month-by-month row counts")
    st.caption("Each bar should be roughly similar height. A very short bar = incomplete scrape for that month.")
    monthly = df_perf.groupby("snapshot_date").size().reset_index(name="rows")
    monthly["month"] = monthly.snapshot_date.dt.strftime("%b %Y")
    avg = monthly.rows.mean()
    monthly["status"] = monthly.rows.apply(lambda x: "✅ Normal" if x >= avg*0.8 else ("⚠️ Low" if x >= avg*0.5 else "❌ Very low"))
    fig = px.bar(monthly, x="month", y="rows", color="status",
                 color_discrete_map={"✅ Normal":"#16a34a","⚠️ Low":"#d97706","❌ Very low":"#dc2626"},
                 title="Row count per snapshot month")
    fig.add_hline(y=avg*0.8, line_dash="dash", line_color="#d97706",
                  annotation_text="80% of avg (warning threshold)")
    fig.update_layout(height=320, template="plotly_white", showlegend=True,
                      xaxis_tickangle=-45)
    st.plotly_chart(fig, use_container_width=True)

    low_months = monthly[monthly.status != "✅ Normal"]
    if not low_months.empty:
        st.markdown("<div class='wb'>⚠️ These months have unusually low row counts — re-scrape may be needed:<br>" +
                    ", ".join(low_months.month.tolist()) + "</div>", unsafe_allow_html=True)
    else:
        st.markdown("<div class='ib'>✅ All months have consistent row counts.</div>", unsafe_allow_html=True)

    # Flag distribution
    st.markdown("### 🚩 Data quality flags")
    fc = df_perf.data_quality_flag.value_counts().reset_index()
    fc.columns = ["Flag","Count"]
    fc["% of total"] = (fc.Count / len(df_perf) * 100).round(2)
    c1,c2 = st.columns(2)
    with c1: st.dataframe(fc, use_container_width=True, hide_index=True)
    with c2:
        fig2 = px.pie(fc, values="Count", names="Flag",
                      color_discrete_sequence=["#16a34a","#94a3b8","#1e40af","#dc2626","#7c3aed","#f59e0b"],
                      title="Flag distribution")
        fig2.update_layout(height=280, template="plotly_white")
        st.plotly_chart(fig2, use_container_width=True)

    clean_pct = (df_perf.data_quality_flag=="CLEAN").mean()*100
    if clean_pct >= 90:
        st.markdown(f"<div class='ib'>✅ {clean_pct:.1f}% of rows are CLEAN — good data quality.</div>", unsafe_allow_html=True)
    elif clean_pct >= 75:
        st.markdown(f"<div class='wb'>⚠️ {clean_pct:.1f}% CLEAN — check flagged rows.</div>", unsafe_allow_html=True)
    else:
        st.markdown(f"<div style='background:#fef2f2;border:1px solid #fecaca;border-radius:8px;padding:12px;color:#991b1b'>❌ Only {clean_pct:.1f}% CLEAN — data quality issue, do not present.</div>", unsafe_allow_html=True)

    # Return range check
    st.markdown("### 📊 Return range sanity check")
    st.caption("1M returns should be between -40% and +40% for CLEAN rows. "
               "Anything outside this range that isn't flagged = potential data error.")
    clean = df_perf[df_perf.data_quality_flag=="CLEAN"]
    if "return_1m" in clean.columns:
        r_min = clean.return_1m.min(); r_max = clean.return_1m.max()
        r_mean = clean.return_1m.mean(); r_std = clean.return_1m.std()
        m1,m2,m3,m4 = st.columns(4)
        m1.metric("Min 1M return", f"{r_min:.2f}%", help="Should be > -40%")
        m2.metric("Max 1M return", f"{r_max:.2f}%", help="Should be < +40%")
        m3.metric("Mean 1M return", f"{r_mean:.2f}%")
        m4.metric("Std Dev 1M", f"{r_std:.2f}%")
        if r_min < -40 or r_max > 40:
            st.markdown("<div class='wb'>⚠️ Some CLEAN rows have returns outside ±40% — review flagging rules.</div>", unsafe_allow_html=True)
        else:
            st.markdown("<div class='ib'>✅ All CLEAN 1M returns within ±40% range.</div>", unsafe_allow_html=True)

    # Benchmark coverage
    st.markdown("### 📐 Benchmark coverage")
    if not df_bench.empty:
        bm = df_bench.groupby("benchmark_name").agg(
            months=("snapshot_date","nunique"),
            has_3y=("return_3y",lambda x: x.notna().sum()),
            has_5y=("return_5y",lambda x: x.notna().sum()),
            latest=("snapshot_date","max")
        ).reset_index()
        bm["latest"] = bm["latest"].dt.strftime("%b %Y")
        bm["status"] = bm.months.apply(lambda x: "✅" if x >= 36 else ("⚠️" if x >= 24 else "❌"))
        st.dataframe(bm.rename(columns={"benchmark_name":"Benchmark","months":"Months",
                                         "has_3y":"3Y rows","has_5y":"5Y rows",
                                         "latest":"Latest","status":"Status"}),
                     use_container_width=True, hide_index=True)

    # Turnover coverage
    st.markdown("### 🔄 Turnover data coverage")
    if not df_turn.empty:
        t_months = df_turn.snapshot_date.nunique()
        t_ias = df_turn.ia_name.nunique()
        t1,t2,t3 = st.columns(3)
        t1.metric("Turnover months", t_months)
        t2.metric("IAs with turnover data", t_ias)
        t3.metric("Latest turnover", df_turn.snapshot_date.max().strftime("%b %Y"))
        coverage = t_ias / df_perf.ia_name.nunique() * 100
        if coverage >= 80:
            st.markdown(f"<div class='ib'>✅ Turnover data covers {coverage:.0f}% of IAs.</div>", unsafe_allow_html=True)
        else:
            st.markdown(f"<div class='wb'>⚠️ Turnover data covers only {coverage:.0f}% of IAs.</div>", unsafe_allow_html=True)

    # Duplicate check
    st.markdown("### 🔁 Duplicate check")
    dupes = df_perf.groupby(["snapshot_date","pms_provider","ia_name","service_type"]).size()
    dupes = dupes[dupes > 1]
    if len(dupes) == 0:
        st.markdown("<div class='ib'>✅ Zero duplicates detected.</div>", unsafe_allow_html=True)
    else:
        st.markdown(f"<div class='wb'>⚠️ {len(dupes)} duplicate rows detected — run dedup before presenting.</div>", unsafe_allow_html=True)
        st.dataframe(dupes.reset_index().rename(columns={0:"count"}).head(20),
                     use_container_width=True, hide_index=True)

    # Late reporters — providers missing from latest snapshot
    st.markdown("### ⏰ Late reporters — missing from latest snapshot")
    st.caption("These providers have historical data but no rows in the most recent snapshot. They may have submitted late to APMI.")
    latest_snap = df_perf.snapshot_date.max()
    prev_snap   = sorted(df_perf.snapshot_date.unique())[-2]
    providers_latest = set(df_perf[df_perf.snapshot_date==latest_snap].pms_provider.unique())
    providers_prev   = set(df_perf[df_perf.snapshot_date==prev_snap].pms_provider.unique())
    missing = sorted(providers_prev - providers_latest)
    if missing:
        st.markdown(f"<div class='wb'>⚠️ {len(missing)} providers reported in {prev_snap.strftime('%b %Y')} but not in {latest_snap.strftime('%b %Y')}:</div>", unsafe_allow_html=True)
        st.dataframe(pd.DataFrame({"Provider": missing}), use_container_width=True, hide_index=True)
        st.caption("Switch sidebar to previous month to see their data.")
    else:
        st.markdown("<div class='ib'>✅ All providers from last month reported in latest snapshot.</div>", unsafe_allow_html=True)

    st.markdown("---")
    st.caption(f"Validation run at {datetime.now().strftime('%Y-%m-%d %H:%M')} · DB: {DB}")




# ── TAB: SCRIPBOX SPOTLIGHT ───────────────────────────────────────────────────
def t_scripbox(df_all, signals, f, db):
    sec("🏦 Scripbox — How Are We Doing?")
    st.markdown("""<div class='ib'>Each Scripbox strategy is ranked within its own category —
    Equity vs Equity, Debt vs Debt, Hybrid vs Hybrid. Scroll down to see the comparison
    with the top 10 peers in each category.</div>""", unsafe_allow_html=True)

    SCRIPBOX = "Wealth Managers (India) Pvt. Ltd."
    p, pl = f["period"], f["period_label"]
    snap_date = f.get("snapshot_date")
    is_cagr = p in CAGR

    full = perf_raw()
    full = snap(full, snap_date)
    full = full[full.data_quality_flag == "CLEAN"]
    if p not in full.columns:
        st.warning(f"{pl} not available."); return

    scripbox_ias = full[full.pms_provider == SCRIPBOX].copy()
    if scripbox_ias.empty:
        st.warning("No Scripbox data for this snapshot. Try switching to March 2026 in the sidebar.")
        return

    snap_label = pd.Timestamp(snap_date).strftime("%B %Y") if snap_date else full.snapshot_date.max().strftime("%B %Y")
    st.markdown(f"**Snapshot:** {snap_label} &nbsp;|&nbsp; **{len(scripbox_ias)} Scripbox strategies across {scripbox_ias.strategy_type.nunique()} categories**")

    # Strategy selector
    strategy_types = sorted(scripbox_ias["strategy_type"].dropna().unique().tolist())
    sel_strat = st.radio("Select Strategy Category", strategy_types, horizontal=True, key="sc_strat")
    st.markdown("---")

    # Filter to selected strategy
    sc_strat = scripbox_ias[scripbox_ias.strategy_type == sel_strat].copy()
    peers = full[full.strategy_type == sel_strat].dropna(subset=[p]).copy()
    peers = peers.sort_values(p, ascending=False).reset_index(drop=True)
    peers["_rank"] = peers.index + 1
    total = len(peers)

    q_colors = {
        "Q1":"#16a34a","Q2":"#86efac","Q3":"#fb923c","Q4":"#dc2626","—":"#94a3b8"
    }
    q_labels_full = {
        "Q1":"Top 25%","Q2":"Top Half","Q3":"Bottom Half","Q4":"Bottom 25%"
    }

    def get_quartile(val):
        if pd.isna(val) or len(peers)==0: return "—"
        pct = (peers[p] <= val).sum() / len(peers)
        return "Q1" if pct>=0.75 else ("Q2" if pct>=0.50 else ("Q3" if pct>=0.25 else "Q4"))

    # ── SUMMARY ROW: max 4 cards per row ─────────────────────────────────────
    st.markdown(f"### {sel_strat} Strategies — Quick Overview")
    cards_per_row = min(4, len(sc_strat))
    rows_needed = (len(sc_strat) + cards_per_row - 1) // cards_per_row

    sc_list = list(sc_strat.iterrows())
    for row_idx in range(rows_needed):
        row_items = sc_list[row_idx*cards_per_row:(row_idx+1)*cards_per_row]
        cols = st.columns(len(row_items))
        for col_idx, (_, row) in enumerate(row_items):
            ret_val = row.get(p)
            rank_val = int(peers[peers.ia_name==row.ia_name]["_rank"].iloc[0]) if row.ia_name in peers.ia_name.values else None
            q = get_quartile(ret_val)
            q_col = q_colors.get(q,"#94a3b8")
            q_lbl = q_labels_full.get(q,"—")
            pct_beat = round((1-rank_val/total)*100) if rank_val else 0
            ret_str = fp(ret_val, is_cagr)
            sig_row = signals[(signals.pms_provider==SCRIPBOX)&(signals.ia_name==row.ia_name)] if not signals.empty else pd.DataFrame()
            sig = sig_row.iloc[0]["signal"] if not sig_row.empty else "—"
            score = f"{int(sig_row.iloc[0].top_half_months)}/{int(sig_row.iloc[0].months_available)}" if not sig_row.empty else "—"
            with cols[col_idx]:
                st.markdown(f"""
<div style='border:2px solid {q_col};border-radius:10px;padding:12px;text-align:center;
background:white;min-height:180px;display:flex;flex-direction:column;justify-content:center'>
<div style='font-size:10px;color:#64748b;font-weight:600;margin-bottom:4px'>{row.ia_name[:30]}</div>
<div style='font-size:24px;font-weight:800;color:{q_col}'>{ret_str}</div>
<div style='font-size:12px;font-weight:700;color:{q_col};margin:2px 0'>{q} — {q_lbl}</div>
<div style='font-size:10px;color:#64748b'>#{rank_val} of {total} peers</div>
<div style='font-size:10px;color:#64748b'>Better than {pct_beat}%</div>
<div style='font-size:10px;color:#64748b;margin-top:4px'>{sig} · {score}</div>
</div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── SIDE BY SIDE: Scripbox vs Top 10 ─────────────────────────────────────
    st.markdown(f"### {sel_strat} — Scripbox vs Top 10 Peers")
    st.caption(f"Left: Top 10 {sel_strat} PMSs by {pl} return. Right: Where each Scripbox strategy sits. Blue = Scripbox, grey = peers, gold line = category median.")

    # Build chart data
    top10 = peers.head(10).copy()
    chart_rows = []
    for _, r in top10.iterrows():
        is_sc = r.ia_name in sc_strat.ia_name.values
        chart_rows.append({
            "Name": ("🏦 " + r.ia_name[:28]) if is_sc else r.ia_name[:28],
            "Return": float(r[p]) if pd.notna(r[p]) else 0,
            "Type": "Scripbox" if is_sc else "Peer",
            "Rank": int(r["_rank"])
        })

    # Add Scripbox strategies not in top 10
    for _, r in sc_strat.iterrows():
        if r.ia_name not in top10.ia_name.values and pd.notna(r.get(p)):
            rank_val2 = int(peers[peers.ia_name==r.ia_name]["_rank"].iloc[0]) if r.ia_name in peers.ia_name.values else None
            chart_rows.append({
                "Name": "🏦 " + r.ia_name[:28],
                "Return": float(r[p]),
                "Type": "Scripbox",
                "Rank": rank_val2 or 999
            })

    if chart_rows:
        df_chart = pd.DataFrame(chart_rows).sort_values("Return", ascending=True)
        cat_median = float(peers[p].median())
        fig = px.bar(df_chart, y="Name", x="Return",
                     color="Type",
                     color_discrete_map={"Scripbox":"#1e40af","Peer":"#cbd5e1"},
                     orientation="h",
                     text=df_chart["Return"].apply(lambda x: f"{x:+.1f}%"),
                     labels={"Return":f"{pl} Return (%)","Name":""},
                     title=f"Scripbox vs Top {min(10,len(top10))} {sel_strat} PMSs — {pl}")
        fig.add_vline(x=cat_median, line_dash="dash", line_color="#f59e0b",
                      annotation_text=f"Median: {cat_median:.1f}%",
                      annotation_font_color="#f59e0b")
        fig.update_traces(textposition="outside")
        fig.update_layout(height=max(350, len(chart_rows)*40),
                          template="plotly_white",
                          legend=dict(orientation="h",y=-0.15),
                          xaxis_title=f"{pl} Return (%)")
        st.plotly_chart(fig, use_container_width=True)

    # ── DETAILED TABLE ────────────────────────────────────────────────────────
    st.markdown(f"### Full Rankings — {sel_strat}")
    st.caption("Showing top 10 peers + all Scripbox strategies. 🏦 = Scripbox strategy.")

    table_rows = []
    for _, r in top10.iterrows():
        is_sc = r.ia_name in sc_strat.ia_name.values
        table_rows.append({
            "Rank": int(r["_rank"]),
            "": "🏦" if is_sc else "",
            "Provider": "Scripbox (WM India)" if is_sc else r.pms_provider[:35],
            "Strategy": r.ia_name[:40],
            f"Return ({pl})": fp(r[p], is_cagr),
            "AUM (₹Cr)": round(r.aum_cr,1) if pd.notna(r.get("aum_cr")) else "—",
        })

    # Add separator + Scripbox not in top 10
    sc_outside = [(r, int(peers[peers.ia_name==r.ia_name]["_rank"].iloc[0]))
                  for _,r in sc_strat.iterrows()
                  if r.ia_name not in top10.ia_name.values and r.ia_name in peers.ia_name.values]
    if sc_outside:
        table_rows.append({"Rank":"·····","":"","Provider":"·····",
                           "Strategy":f"({sc_outside[0][1]-11} more funds)","Return ({pl})":"","AUM (₹Cr)":""})
        for r, rank3 in sorted(sc_outside, key=lambda x: x[1]):
            table_rows.append({
                "Rank": rank3,
                "": "🏦",
                "Provider": "Scripbox (WM India)",
                "Strategy": r.ia_name[:40],
                f"Return ({pl})": fp(r.get(p), is_cagr),
                "AUM (₹Cr)": round(r.aum_cr,1) if pd.notna(r.get("aum_cr")) else "—",
            })

    st.dataframe(pd.DataFrame(table_rows), use_container_width=True, hide_index=True)

    # ── PLAIN ENGLISH VERDICT ─────────────────────────────────────────────────
    st.markdown("### What does this mean?")
    for _, row in sc_strat.iterrows():
        ret_val = row.get(p)
        rank_val4 = int(peers[peers.ia_name==row.ia_name]["_rank"].iloc[0]) if row.ia_name in peers.ia_name.values else None
        q = get_quartile(ret_val)
        q_col = q_colors.get(q,"#64748b")
        pct_beat3 = round((1-rank_val4/total)*100) if rank_val4 else 0
        gap_to_median = float(ret_val)-float(peers[p].median()) if pd.notna(ret_val) else None
        gap_to_top = float(ret_val)-float(peers[p].iloc[0]) if pd.notna(ret_val) else None
        sig_row2 = signals[(signals.pms_provider==SCRIPBOX)&(signals.ia_name==row.ia_name)] if not signals.empty else pd.DataFrame()
        sig2 = sig_row2.iloc[0]["signal"] if not sig_row2.empty else "—"

        if "Q1" in q:
            msg = f"✅ <b>{row.ia_name}</b> is performing in the <b>top 25%</b> of all {sel_strat} PMSs — better than {pct_beat3}% of peers. It is {fp(abs(gap_to_median) if gap_to_median else 0, False)} ahead of the category average. Consistency signal: <b>{sig2}</b>."
        elif "Q2" in q:
            msg = f"🟡 <b>{row.ia_name}</b> is in the <b>top half</b> of {sel_strat} PMSs but not yet top 25%. It is {fp(gap_to_median, False)} vs the category average, and {fp(gap_to_top, False)} behind the top fund. Consistency signal: <b>{sig2}</b>."
        elif "Q3" in q:
            msg = f"🟠 <b>{row.ia_name}</b> is in the <b>bottom half</b> of {sel_strat} PMSs — {fp(abs(gap_to_median) if gap_to_median else 0, False)} below the category average. Consistency signal: <b>{sig2}</b>."
        else:
            msg = f"🔴 <b>{row.ia_name}</b> is in the <b>bottom 25%</b> of {sel_strat} PMSs. It is significantly behind category peers. Consistency signal: <b>{sig2}</b>."

        st.markdown(f"<div style='border-left:4px solid {q_col};padding:10px 14px;margin:6px 0;border-radius:0 8px 8px 0;font-size:13px;background:#f8fafc'>{msg}</div>",
                    unsafe_allow_html=True)

    # ── OVERALL SUMMARY TABLE ─────────────────────────────────────────────────
    with st.expander("📊 All Scripbox strategies — full summary"):
        summary = []
        for _, row in scripbox_ias.iterrows():
            st2 = row.get("strategy_type","")
            peers2 = full[full.strategy_type==st2].dropna(subset=[p]).sort_values(p,ascending=False).reset_index(drop=True)
            peers2["_rank"] = peers2.index+1
            rank2 = int(peers2[peers2.ia_name==row.ia_name]["_rank"].iloc[0]) if row.ia_name in peers2.ia_name.values else None
            total2 = len(peers2)
            q2 = "Q1" if rank2 and rank2/total2<=0.25 else ("Q2" if rank2 and rank2/total2<=0.50 else ("Q3" if rank2 and rank2/total2<=0.75 else "Q4")) if rank2 else "—"
            sig_r = signals[(signals.pms_provider==SCRIPBOX)&(signals.ia_name==row.ia_name)] if not signals.empty else pd.DataFrame()
            summary.append({
                "Strategy": row.ia_name[:40],
                "Category": st2,
                f"Return ({pl})": fp(row.get(p), is_cagr),
                "Rank": f"#{rank2} of {total2}" if rank2 else "—",
                "Quartile": q2,
                "Signal": sig_r.iloc[0]["signal"] if not sig_r.empty else "—",
                "Score": f"{int(sig_r.iloc[0].top_half_months)}/{int(sig_r.iloc[0].months_available)}" if not sig_r.empty else "—",
            })
        st.dataframe(pd.DataFrame(summary), use_container_width=True, hide_index=True)
        csvb(pd.DataFrame(summary), "scripbox_all_strategies.csv", "⬇️ CSV — All Scripbox Strategies")


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    pw = perf_with_bench(); pr = perf_raw(); db = bench_data()
    dr = risk_data(); da = aum_decomp(); dt = turnover_data()

    if pw.empty: st.error("⚠️ DB empty. Copy apmi_pms.db to /tmp/ first."); return

    with st.spinner("Computing 18-month consistency signals..."):
        signals = compute_signals()

    f = sidebar(pw)
    df = apply_filters(pw.copy(), f, signals)

    st.markdown("""<div style="padding:20px 0 8px">
    <h1 style="color:#1e40af;margin:0;font-size:30px">📊 APMI PMS Performance Dashboard</h1>
    <p style="color:#64748b;margin:4px 0 0;font-size:14px">SEBI-mandated TWRR · All registered PMS Investment Approaches · Scripbox</p>
    </div>""", unsafe_allow_html=True)

    if not df.empty:
        k1,k2,k3,k4,k5 = st.columns(5)
        k1.metric("IAs", f"{df.ia_name.nunique():,}")
        k2.metric("Providers", f"{df.pms_provider.nunique():,}")
        k3.metric("Total AUM", f"₹{df.aum_cr.sum():,.0f} Cr")
        a1 = df.return_1y.mean() if "return_1y" in df.columns else None
        k4.metric("Avg 1Y Return", fp(a1, True) if pd.notna(a1) else "—")
        k5.metric("Latest", pw.snapshot_date.max().strftime("%b %Y"))

    # ── MORNING BRIEFING ──────────────────────────────────────────────────────
    with st.expander("☀️ Morning Briefing — open this first", expanded=True):
        if not signals.empty and not df.empty:
            strats = f.get("strategies",[])
            sig_sub = signals[signals.strategy_type.isin(strats)] if strats else signals
            snap_date = f.get("snapshot_date")
            df_snap = snap(perf_raw(), snap_date)
            df_clean_snap = df_snap[df_snap.data_quality_flag=="CLEAN"]
            db_snap = snap(bench_data(), snap_date)
            nifty_snap = db_snap[db_snap.benchmark_name.str.contains("Nifty 50",na=False)]
            nifty_1m = float(nifty_snap.return_1m.iloc[0]) if not nifty_snap.empty and pd.notna(nifty_snap.return_1m.iloc[0]) else None
            nifty_1y = float(nifty_snap.return_1y.iloc[0]) if not nifty_snap.empty and pd.notna(nifty_snap.return_1y.iloc[0]) else None

            rec_ias = sig_sub[sig_sub.signal=="Recommended"]
            n_rec = len(rec_ias); n_hold = (sig_sub.signal=="Hold").sum(); n_nr = (sig_sub.signal=="Not Recommended").sum()

            # Q1 — Is this PMS doing what it's supposed to do?
            beat_nifty_1m = beat_nifty_1y = None
            if nifty_1m is not None and "return_1m" in df_clean_snap.columns:
                beat_nifty_1m = (df_clean_snap.return_1m > nifty_1m).sum()
                total_1m = df_clean_snap.return_1m.notna().sum()
            if nifty_1y is not None and "return_1y" in df_clean_snap.columns:
                beat_nifty_1y = (df_clean_snap.return_1y > nifty_1y).sum()
                total_1y = df_clean_snap.return_1y.notna().sum()
            med_1m = df_clean_snap.return_1m.median() if "return_1m" in df_clean_snap.columns else None
            med_1y = df_clean_snap.return_1y.median() if "return_1y" in df_clean_snap.columns else None

            # Q2 — Consistency
            pct_rec = round(n_rec / len(sig_sub) * 100, 1) if len(sig_sub) > 0 else 0

            # Q3 — Red flags
            _, rec_df2 = load_recommendations() if False else (None, pd.DataFrame())
            rec_from_sig = sig_sub[sig_sub.signal=="Recommended"]
            flag_count = 0
            if not rec_from_sig.empty:
                for _, r2 in rec_from_sig.iterrows():
                    ia2 = df_clean_snap[(df_clean_snap.pms_provider==r2.pms_provider)&(df_clean_snap.ia_name==r2.ia_name)]
                    if not ia2.empty and nifty_1m is not None:
                        if pd.notna(ia2.return_1m.iloc[0]) and float(ia2.return_1m.iloc[0]) < nifty_1m:
                            flag_count += 1

            # Trend alerts count
            trend_flag_count = 0
            all_dates_b = sorted(perf_raw()["snapshot_date"].unique(), reverse=True)
            recent_dates = all_dates_b[:6]; prev_dates = all_dates_b[6:12]
            if len(prev_dates) >= 3:
                nifty_m_b = (bench_data()[bench_data().benchmark_name.str.contains("Nifty 50",na=False)]
                             [["snapshot_date","return_1m"]].rename(columns={"return_1m":"nifty_1m"}))
                for _, r2 in rec_from_sig.head(20).iterrows():
                    ia_h = perf_raw()[(perf_raw().pms_provider==r2.pms_provider)&
                                      (perf_raw().ia_name==r2.ia_name)&
                                      (perf_raw().data_quality_flag=="CLEAN")].copy()
                    ia_h = ia_h.merge(nifty_m_b, on="snapshot_date", how="left")
                    ia_h["alpha_1m"] = ia_h["return_1m"] - ia_h["nifty_1m"].fillna(0)
                    strat_b = r2.strategy_type
                    def quick_score(dates):
                        sub_b = ia_h[ia_h.snapshot_date.isin(dates)]
                        if len(sub_b) < 3: return None
                        hits_b = 0
                        for _, row_b in sub_b.iterrows():
                            peers_b = perf_raw()[(perf_raw().snapshot_date==row_b.snapshot_date)&
                                                  (perf_raw().strategy_type==strat_b)&
                                                  (perf_raw().data_quality_flag=="CLEAN")]
                            peers_b = peers_b.merge(nifty_m_b, on="snapshot_date", how="left")
                            peers_b["alpha_1m"] = peers_b["return_1m"] - peers_b["nifty_1m"].fillna(0)
                            ap = (peers_b["return_1m"].dropna()<=row_b.return_1m).sum()/max(len(peers_b["return_1m"].dropna()),1) if pd.notna(row_b.get("return_1m")) else 0
                            alp = (peers_b["alpha_1m"].dropna()<=row_b.alpha_1m).sum()/max(len(peers_b["alpha_1m"].dropna()),1) if pd.notna(row_b.get("alpha_1m")) else 0
                            if ap >= 0.50 and alp >= 0.50: hits_b += 1
                        return round(hits_b/len(sub_b)*100,1)
                    sc_r = quick_score(recent_dates); sc_p = quick_score(prev_dates)
                    if sc_r is not None and sc_p is not None and (sc_p - sc_r) >= 20:
                        trend_flag_count += 1

            snap_label = pd.Timestamp(snap_date).strftime("%B %Y") if snap_date else perf_raw().snapshot_date.max().strftime("%B %Y")
            st.markdown(f"**As of {snap_label}** — {df_clean_snap.ia_name.nunique():,} IAs · {df_clean_snap.pms_provider.nunique():,} providers")
            st.markdown("---")
            c1, c2, c3 = st.columns(3)
            with c1:
                st.markdown("**❓ Are PMSs doing what they should?**")
                if beat_nifty_1m is not None:
                    color = "#16a34a" if beat_nifty_1m/total_1m > 0.5 else "#dc2626"
                    st.markdown(f"<span style='color:{color};font-size:18px;font-weight:700'>{beat_nifty_1m}/{total_1m} IAs beat Nifty 50 this month</span>", unsafe_allow_html=True)
                if beat_nifty_1y is not None:
                    color2 = "#16a34a" if beat_nifty_1y/total_1y > 0.5 else "#dc2626"
                    st.markdown(f"<span style='color:{color2};font-size:15px'>{beat_nifty_1y}/{total_1y} beat Nifty 50 on 1Y</span>", unsafe_allow_html=True)
                if med_1m is not None:
                    nifty_str = f" vs Nifty {nifty_1m:+.2f}%" if nifty_1m else ""
                    st.markdown(f"Median 1M return: **{med_1m:+.2f}%**{nifty_str}")
                if med_1y is not None:
                    st.markdown(f"Median 1Y return: **{med_1y:+.2f}% CAGR**")
            with c2:
                st.markdown("**📊 Is performance consistent?**")
                color3 = "#16a34a" if pct_rec >= 5 else "#d97706"
                st.markdown(f"<span style='color:{color3};font-size:18px;font-weight:700'>{n_rec} Recommended</span> · {n_hold} Hold · {n_nr} Not Recommended", unsafe_allow_html=True)
                st.markdown(f"Only **{pct_rec}%** of all IAs meet the 18-month consistency bar")
                if not rec_from_sig.empty:
                    top3 = rec_from_sig.sort_values("consistency_score", ascending=False).head(3)
                    st.markdown("**Top 3 most consistent:**")
                    for _, t in top3.iterrows():
                        st.markdown(f"• {t.ia_name[:35]} ({int(t.top_half_months)}/{int(t.months_available)})")
            with c3:
                st.markdown("**🚨 Red flags to act on?**")
                total_flags = flag_count + trend_flag_count
                if total_flags == 0:
                    st.markdown("<span style='color:#16a34a;font-size:18px;font-weight:700'>✅ No red flags today</span>", unsafe_allow_html=True)
                else:
                    st.markdown(f"<span style='color:#dc2626;font-size:18px;font-weight:700'>⚠️ {total_flags} items need attention</span>", unsafe_allow_html=True)
                if flag_count > 0:
                    st.markdown(f"• **{flag_count} Recommended IA(s)** underperforming Nifty 50 this month → check Flags tab")
                if trend_flag_count > 0:
                    st.markdown(f"• **{trend_flag_count} Recommended IA(s)** showing consistency decline → check Flags tab")
                if total_flags == 0:
                    st.markdown("All Recommended IAs outperforming benchmarks")
                    st.markdown("No consistency decline detected")
        else:
            st.info("Loading signals...")

    if f.get("show_flagged"):
        st.markdown("<div class='wb'>⚠️ Flagged data enabled.</div>", unsafe_allow_html=True)
    st.markdown("---")

    tabs = st.tabs([
        "🏆 Leaderboard", "🏦 Scripbox", "🎯 Signals", "📐 Benchmark Compare",
        "📈 NAV Growth", "🔢 Quartile Analysis", "💰 AUM Trend",
        "🚨 Flags & Alerts", "🔀 IA Compare", "📈 Time-Series",
        "⚖️ Risk Metrics", "🔄 Turnover", "🏢 Provider Compare",
        "🔍 Data Validation"
    ])
    with tabs[0]:  t_leaderboard(df, f, db, signals)
    with tabs[1]:  t_scripbox(pw, signals, f, db)
    with tabs[2]:  t_signals(signals, f, pr, db)
    with tabs[3]:  t_bench(pr, f, db, signals)
    with tabs[4]:  t_nav(pr, f, db, signals)
    with tabs[5]:  t_quartile(pw, f, db)
    with tabs[6]:  t_aum(da, f, pr)
    with tabs[7]:  t_flags(pr, db, signals, f)
    with tabs[8]:  t_compare(pr, dr, f, db, signals)
    with tabs[9]:  t_ts(pw, f, db)
    with tabs[10]: t_risk(dr, df, f)
    with tabs[11]: t_turn(dt, pw, f)
    with tabs[12]: t_provider(pw, dr, f, db)
    with tabs[13]: t_validation(pr, db, dt)
    st.markdown("<div style='text-align:center;color:#94a3b8;font-size:11px;padding:12px 0'>Data: APMI · SEBI-mandated TWRR · Not investment advice · Scripbox</div>", unsafe_allow_html=True)

if __name__ == "__main__": main()
